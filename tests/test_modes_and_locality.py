from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest

from openpyxl import Workbook

from safety_limits import InputDirectoryLimits
from smetchik_engine import InspectionInputError, inspect_input
from test_cli_and_result import run_cli


REQUIRED_RESULT_FIELDS = {
    "schema_version",
    "run_id",
    "mode",
    "purpose",
    "execution_status",
    "context",
    "input_inventory",
    "coverage",
    "estimate_hierarchy",
    "checks",
    "amounts",
    "cost_analytics",
    "findings",
    "questions",
    "limitations",
    "normative_sources",
    "artifacts",
    "overall_status",
    "recommended_action",
}


def test_light_and_full_have_distinct_honest_coverage(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["code", "name", "quantity", "unit", "unit_price", "total"])
    sheet.append(["R-1", "Работа", 2, "м2", 10, 20])
    workbook.save(source)
    before = hashlib.sha256(source.read_bytes()).hexdigest()

    context = {
        "object": "Объект",
        "work_type": "construction",
        "funding_source": "budget",
        "region_or_price_zone": "67",
        "price_level_date": "2026-08-01",
        "calculation_method": "resource_index",
        "stage": "project_documentation",
        "document_set": ["LSR"],
    }
    light = inspect_input(source, mode="light", purpose="internal_review", context=context)
    full = inspect_input(source, mode="full", purpose="internal_review", context=context)

    assert light["coverage"]["row_level_checked"] is False
    assert light["coverage"]["sampling_strategy"] == "none"
    assert light["coverage"]["checked_records"] == 0
    assert "процент" not in light["coverage"]["description"].casefold()
    assert full["coverage"]["row_level_checked"] is False
    assert full["coverage"]["checked_records"] == 0
    assert full["coverage"]["arithmetic_checked_records"] == 0
    assert next(check for check in full["checks"] if check["id"] == "ARITH-01")["status"] == "limited"
    assert any(
        limitation["code"] == "full_row_controls_incomplete"
        for limitation in full["limitations"]
    )
    assert full["coverage"]["extracted_records"] >= full["coverage"]["checkable_records"]
    assert set(full) == REQUIRED_RESULT_FIELDS
    assert full["findings"] == []
    assert full["normative_sources"] == []
    assert hashlib.sha256(source.read_bytes()).hexdigest() == before


def test_full_mode_describes_maximal_available_coverage_with_limits(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "=SUM(1,2)"
    workbook.save(source)

    result = inspect_input(source, mode="full", purpose="internal_review", context={
        "object": "Объект",
        "work_type": "construction",
        "funding_source": "budget",
        "region_or_price_zone": "67",
        "price_level_date": "2026-08-01",
        "calculation_method": "resource_index",
        "stage": "project_documentation",
        "document_set": ["LSR"],
    })

    assert result["execution_status"] == "completed_with_limits"
    assert result["overall_status"] == "verification_impossible"
    assert any(
        limitation["code"] == "no_estimate_rows_available"
        for limitation in result["limitations"]
    )
    assert result["coverage"]["status"] == "completed_with_limits"
    assert "максимально сплошная проверка доступных данных с ограничениями" in result["coverage"]["description"].casefold()


def test_cli_reads_context_and_creates_only_a_new_output_file(tmp_path: Path) -> None:
    source = tmp_path / "estimate.bin"
    source.write_bytes(b"local")
    context = tmp_path / "context.json"
    context.write_text('{"document_role":"project_estimate"}', encoding="utf-8")
    output = tmp_path / "result.json"

    completed = run_cli(
        "inspect",
        source,
        "--mode",
        "light",
        "--context-json",
        context,
        "--output-json",
        output,
    )

    assert completed.returncode == 0, completed.stderr
    assert completed.stdout == ""
    result = json.loads(output.read_text(encoding="utf-8"))
    assert result["context"] == {"document_role": "project_estimate"}
    assert result["execution_status"] == "needs_input"


def test_invalid_context_fails_without_sensitive_traceback(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xml"
    source.write_text("<root/>", encoding="utf-8")
    context = tmp_path / "context.json"
    context.write_text("not-json SECRET_CONTENT", encoding="utf-8")

    completed = run_cli(
        "inspect",
        source,
        "--mode",
        "full",
        "--context-json",
        context,
    )

    combined = completed.stdout + completed.stderr
    assert completed.returncode == 2
    assert "traceback" not in combined.casefold()
    assert "SECRET_CONTENT" not in combined


def test_supported_local_extraction_never_opens_a_network_connection(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "estimate.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = 1
    workbook.save(source)

    def network_forbidden(*_args, **_kwargs):
        raise AssertionError("network access is forbidden")

    monkeypatch.setattr("socket.create_connection", network_forbidden)

    result = inspect_input(source, mode="full", purpose="internal_review", context={
        "object": "Объект",
        "work_type": "construction",
        "funding_source": "budget",
        "region_or_price_zone": "67",
        "price_level_date": "2026-08-01",
        "calculation_method": "resource_index",
        "stage": "project_documentation",
        "document_set": ["LSR"],
    })

    assert result["execution_status"] == "completed_with_limits"


def test_directory_symlink_is_rejected_before_target_is_read(tmp_path: Path) -> None:
    secret = tmp_path / "outside-secret.xml"
    secret.write_text("<secret>DO_NOT_READ</secret>", encoding="utf-8")
    package = tmp_path / "package"
    package.mkdir()
    (package / "linked.xml").symlink_to(secret)

    result = inspect_input(package, mode="full")

    assert result["input_inventory"] == []
    assert any(limit["code"] == "input_symlink_forbidden" for limit in result["limitations"])
    assert "DO_NOT_READ" not in json.dumps(result, ensure_ascii=False)


@pytest.mark.parametrize(
    ("limits", "payloads"),
    [
        (InputDirectoryLimits(max_files=1, max_total_bytes=1_000, max_depth=8), [b"a", b"b"]),
        (InputDirectoryLimits(max_files=10, max_total_bytes=3, max_depth=8), [b"four"]),
    ],
)
def test_directory_limits_are_checked_before_any_file_content_is_read(
    tmp_path: Path,
    monkeypatch,
    limits: InputDirectoryLimits,
    payloads: list[bytes],
) -> None:
    package = tmp_path / "package"
    package.mkdir()
    for index, payload in enumerate(payloads):
        (package / f"{index}.bin").write_bytes(payload)
    monkeypatch.setattr("smetchik_engine.INPUT_DIRECTORY_LIMITS", limits)

    def content_read_forbidden(_path: Path) -> str:
        raise AssertionError("preflight must reject before hashing")

    monkeypatch.setattr("smetchik_engine.sha256_file", content_read_forbidden)

    result = inspect_input(package, mode="light")

    assert result["input_inventory"] == []
    assert any(limit["code"] == "input_directory_limit_exceeded" for limit in result["limitations"])


def test_single_file_size_limit_is_checked_before_hashing(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "oversized.bin"
    source.write_bytes(b"four")
    monkeypatch.setattr(
        "smetchik_engine.INPUT_DIRECTORY_LIMITS",
        InputDirectoryLimits(max_files=10, max_total_bytes=3, max_depth=8),
    )

    def content_read_forbidden(_path: Path) -> str:
        raise AssertionError("preflight must reject before hashing")

    monkeypatch.setattr("smetchik_engine.sha256_file", content_read_forbidden)

    result = inspect_input(source, mode="light")

    assert result["input_inventory"] == []
    assert any(limit["code"] == "input_file_limit_exceeded" for limit in result["limitations"])


def test_public_context_cannot_override_registry_or_inject_records(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xml"
    source.write_text("<root/>", encoding="utf-8")

    for reserved in (
        "xml_schema_registry",
        "semantic_records",
        "canonical_records",
        "full_row_coverage",
    ):
        with pytest.raises(InspectionInputError):
            inspect_input(source, mode="full", context={reserved: []})


def test_production_engine_calls_xml_extractor_without_registry_override(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "estimate.xml"
    source.write_text("<root/>", encoding="utf-8")
    captured: dict[str, object] = {}

    def extractor(_path: Path, **kwargs: object):
        captured.update(kwargs)
        return ({"format": "xml", "schema": None, "schema_validation": {"status": "unsupported_schema"}, "structure": [], "semantic_interpretation_performed": False, "arithmetic_performed": False}, "partial", [], 0)

    monkeypatch.setattr("smetchik_engine.extract_xml", extractor)

    inspect_input(source, mode="light")

    assert captured == {}


@pytest.mark.parametrize(
    ("mode", "purpose", "context"),
    [
        ("automatic", None, {}),
        ("light", 123, {}),
        ("full", "x" * 300, {}),
        ("full", None, []),
        ("full", None, {"document_set": [object()]}),
    ],
)
def test_engine_rejects_invalid_mode_purpose_and_context_safely(
    tmp_path: Path,
    mode: object,
    purpose: object,
    context: object,
) -> None:
    source = tmp_path / "estimate.xml"
    source.write_text("<root/>", encoding="utf-8")

    with pytest.raises(InspectionInputError):
        inspect_input(source, mode=mode, purpose=purpose, context=context)  # type: ignore[arg-type]


def test_cli_rejects_reserved_context_without_traceback(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xml"
    source.write_text("<root/>", encoding="utf-8")
    context = tmp_path / "context.json"
    context.write_text('{"semantic_records": []}', encoding="utf-8")

    completed = run_cli("inspect", source, "--mode", "full", "--context-json", context)

    assert completed.returncode == 2
    assert "traceback" not in (completed.stdout + completed.stderr).casefold()
    assert "inspection failed" not in completed.stderr.casefold()


def test_top_level_extractor_evidence_does_not_leak_absolute_input_path(tmp_path: Path) -> None:
    source = tmp_path / "unknown.xml"
    source.write_text("<Vendor><Total>1</Total></Vendor>", encoding="utf-8")

    result = inspect_input(source, mode="light")

    evidence = next(
        limitation["evidence"][0]
        for limitation in result["limitations"]
        if limitation["code"] == "unsupported_schema"
    )
    assert evidence["source_path"] == "unknown.xml"
    assert evidence["locator"].startswith("unknown.xml")
    assert str(tmp_path) not in json.dumps(result, ensure_ascii=False)


def test_missing_passport_blocks_domain_findings_and_passed_checks(tmp_path: Path) -> None:
    source = tmp_path / "estimate.bin"
    source.write_bytes(b"fixture")
    row = {
        "entity": "estimate_row",
        "row_id": "r-1",
        "name": "Работа",
        "quantity": "2",
        "unit": "шт",
        "unit_price": "10",
        "declared_total": "999",
        "evidence": {
            "source_path": "estimate.xlsx",
            "sheet": "ЛСР",
            "cell_range": "A2:F2",
            "locator": "ЛСР!A2:F2",
        },
    }

    result = inspect_input(source, mode="full", _trusted_records=[row])

    assert result["execution_status"] == "needs_input"
    assert len(result["questions"]) == 1
    assert result["findings"] == []
    assert result["coverage"]["checked_records"] == 0
    assert result["coverage"]["row_level_checked"] is False
    assert all(check["status"] not in {"passed", "finding"} for check in result["checks"])


def test_missing_passport_does_not_invoke_substantive_domain_checks(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "estimate.bin"
    source.write_bytes(b"fixture")

    def substantive_check_forbidden(*_args, **_kwargs):
        raise AssertionError("substantive checks must wait for the complete passport")

    monkeypatch.setattr("smetchik_engine.run_domain_checks", substantive_check_forbidden)

    result = inspect_input(source, mode="full")

    assert result["execution_status"] == "needs_input"
    assert len(result["questions"]) == 1


def test_public_inventory_is_compact_and_excludes_raw_extractor_payloads(
    tmp_path: Path,
    monkeypatch,
) -> None:
    package = tmp_path / "package"
    package.mkdir()
    for name in ("estimate.xlsx", "estimate.pdf", "estimate.xml"):
        (package / name).write_bytes(b"fixture")
    secrets = {
        "excel": "EXCEL_UNIQUE_SECRET_92341",
        "pdf": "PDF_UNIQUE_SECRET_92341",
        "xml": "XML_UNIQUE_SECRET_92341",
    }

    def excel_extractor(_path: Path):
        return (
            {
                "format": "xlsx",
                "sheets": [{"name": "ЛСР", "state": "visible", "max_row": 2, "max_column": 6, "cells": [{"coordinate": "A2", "value": secrets["excel"], "formula": "=SECRET_FORMULA"}], "semantic_tables": []}],
                "semantic_row_count": 0,
                "formula_count": 1,
                "macro_capable": False,
                "macros_executed": False,
            },
            "reliable",
            [],
            1,
        )

    def pdf_extractor(_path: Path):
        return (
            {
                "format": "pdf",
                "page_count": 1,
                "pages": [{"page": 1, "rotation": 0, "text": secrets["pdf"], "words": [{"text": secrets["pdf"], "bbox": [1, 2, 3, 4]}], "ocr_words": [], "tables": [], "record_count": 1, "extraction_status": "reliable", "visual_verification_required": False}],
                "visual_verification_required": False,
            },
            "reliable",
            [],
            1,
        )

    def xml_extractor(_path: Path):
        return (
            {
                "format": "xml",
                "schema": {"id": "minstroy.local_estimate_rim.3_01", "version": "3.01"},
                "schema_validation": {"status": "valid", "error_count": 0},
                "structure": [{"local_name": secrets["xml"], "evidence": {"source_path": str(_path), "xpath": "/Construction", "line": 1, "locator": f"{_path}:xpath:/Construction;line:1"}}],
                "semantic_values": [{"local_name": "Name", "raw_value": secrets["xml"], "xpath": "/Construction/Name"}],
                "semantic_records": [],
                "semantic_interpretation_performed": True,
                "arithmetic_performed": False,
            },
            "reliable",
            [],
            1,
        )

    monkeypatch.setattr("smetchik_engine.extract_xlsx", excel_extractor)
    monkeypatch.setattr("smetchik_engine.extract_pdf", pdf_extractor)
    monkeypatch.setattr("smetchik_engine.extract_xml", xml_extractor)

    result = inspect_input(package, mode="light")
    serialized = json.dumps(result, ensure_ascii=False)

    assert all(secret not in serialized for secret in secrets.values())
    assert "SECRET_FORMULA" not in serialized
    assert result["coverage"]["extractor_records"] == 3
    xml_item = next(item for item in result["input_inventory"] if item["file_type"] == "xml")
    assert xml_item["details"]["schema"]["id"] == "minstroy.local_estimate_rim.3_01"
    assert xml_item["details"]["schema_validation"]["status"] == "valid"

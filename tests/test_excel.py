from __future__ import annotations

import hashlib
import struct
from pathlib import Path

from openpyxl import Workbook

from excel_extract import extract_xls, extract_xlsx
from smetchik_engine import inspect_input


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def inventory_item(result: dict, suffix: str) -> dict:
    return next(item for item in result["input_inventory"] if item["path"].endswith(suffix))


def biff_record(record_id: int, payload: bytes = b"") -> bytes:
    return struct.pack("<HH", record_id, len(payload)) + payload


def make_minimal_biff8_xls(path: Path) -> None:
    bof_global = biff_record(
        0x0809,
        struct.pack("<HHHHII", 0x0600, 0x0005, 0x0DBB, 0x07CC, 0x41, 0x06),
    )
    codepage = biff_record(0x0042, struct.pack("<H", 1200))
    name = b"Sheet1"
    bounds_length = 4 + 2 + 1 + 1 + len(name)
    placeholder = biff_record(0x0085, b"\0" * bounds_length)
    eof = biff_record(0x000A)
    sheet_offset = len(bof_global) + len(codepage) + len(placeholder) + len(eof)
    bounds = biff_record(
        0x0085,
        struct.pack("<IBBB", sheet_offset, 0, 0, len(name)) + b"\0" + name,
    )
    bof_sheet = biff_record(
        0x0809,
        struct.pack("<HHHHII", 0x0600, 0x0010, 0x0DBB, 0x07CC, 0x41, 0x06),
    )
    dimensions = biff_record(0x0200, struct.pack("<IIHHH", 0, 1, 0, 1, 0))
    number = biff_record(0x0203, struct.pack("<HHHd", 0, 0, 0, 42.5))
    path.write_bytes(bof_global + codepage + bounds + eof + bof_sheet + dimensions + number + eof)


def test_xlsx_keeps_formulas_and_cached_values_distinct(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estimate"
    sheet["A1"] = 10
    sheet["B1"] = 5
    sheet["C1"] = "=A1+B1*2"
    sheet["D1"] = "=SUM(A1:B1)"
    sheet["E1"] = "='[external.xlsx]Data'!A1"
    sheet.merge_cells("A3:B3")
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 99
    workbook.save(source)
    before = file_hash(source)

    details, status, limitations, records = extract_xlsx(source)
    estimate = next(sheet for sheet in details["sheets"] if sheet["name"] == "Estimate")
    cells = {cell["coordinate"]: cell for cell in estimate["cells"]}
    assert cells["C1"]["formula"] == "=A1+B1*2"
    assert cells["C1"]["cached_value"] is None
    assert cells["C1"]["computed_safe_value"] == 20
    assert cells["D1"]["computed_safe_value"] == 15
    assert cells["E1"]["computed_safe_value"] is None
    assert cells["C1"]["evidence"]["locator"] == "Estimate!C1"
    assert details["hidden_sheets"] == ["Hidden"]
    assert details["merged_cells"] == ["Estimate!A3:B3"]
    assert details["external_links_detected"] is True
    assert "Estimate!C1" in details["missing_cached_formula_values"]
    assert status == "partial"
    assert records == 3
    assert any(limit["code"] == "unsupported_formula" for limit in limitations)

    result = inspect_input(source, mode="full")
    item = inventory_item(result, ".xlsx")
    summary = item["details"]
    assert summary["sheet_count"] == 2
    assert summary["hidden_sheet_count"] == 1
    assert summary["merged_range_count"] == 1
    assert summary["formula_count"] == 3
    assert "cells" not in summary["sheets"][0]
    assert any(limit["code"] == "unsupported_formula" for limit in result["limitations"])
    assert result["coverage"]["row_level_checked"] is False
    assert result["coverage"]["checked_records"] == 0
    assert result["coverage"]["extractor_records"] == 3
    assert file_hash(source) == before


def test_xlsm_is_flagged_and_macros_are_never_executed(tmp_path: Path) -> None:
    source = tmp_path / "estimate.xlsm"
    workbook = Workbook()
    workbook.active["A1"] = "plain"
    workbook.save(source)
    before = file_hash(source)

    result = inspect_input(source, mode="light")

    details = inventory_item(result, ".xlsm")["details"]
    assert details["macro_capable"] is True
    assert details["macros_executed"] is False
    assert result["coverage"]["row_level_checked"] is False
    assert result["coverage"]["sampling_strategy"] == "none"
    assert file_hash(source) == before


def test_unreadable_xls_becomes_a_limitation_without_mutation(tmp_path: Path) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"not-an-ole-workbook")
    before = file_hash(source)

    result = inspect_input(source, mode="full")

    item = inventory_item(result, ".xls")
    assert item["extraction_status"] == "failed"
    assert any(limit["code"] == "unreadable_legacy_xls" for limit in result["limitations"])
    assert result["execution_status"] == "needs_input"
    assert file_hash(source) == before


def test_legacy_xls_is_read_only_and_returns_cached_values(tmp_path: Path) -> None:
    source = tmp_path / "legacy.xls"
    make_minimal_biff8_xls(source)
    before = file_hash(source)

    details, status, limitations, records = extract_xls(source)
    assert details["read_only"] is True
    assert details["sheets"][0]["cells"][0]["cached_value"] == 42.5
    assert details["sheets"][0]["cells"][0]["evidence"]["locator"] == "Sheet1!R1C1"
    assert status == "partial"
    assert records == 0
    assert any(limit["code"] == "legacy_xls_formula_metadata_unavailable" for limit in limitations)

    result = inspect_input(source, mode="full")
    item = inventory_item(result, ".xls")
    assert item["details"]["read_only"] is True
    assert item["details"]["sheet_count"] == 1
    assert "cells" not in item["details"]["sheets"][0]
    assert file_hash(source) == before

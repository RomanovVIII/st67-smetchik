from __future__ import annotations

import hashlib
import json
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
PYTHON = Path(sys.executable)
CLI = ROOT / "skills" / "smetchik" / "scripts" / "smetchik_cli.py"


def run_cli(*args: object) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [str(PYTHON), str(CLI), *map(str, args)],
        cwd=ROOT,
        text=True,
        capture_output=True,
        check=False,
        timeout=30,
    )


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def make_workbook(path: Path, value: object = 10) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ЛСР"
    sheet["B2"] = value
    workbook.save(path)


def write_corrections(
    path: Path,
    source: Path,
    *,
    status: str = "confirmed",
    expected_old: object = 10,
    replacement: object = 12,
    reason: str = "Подтверждено исходными документами",
    target: dict[str, object] | None = None,
    source_hash: str | None = None,
) -> None:
    path.write_text(
        json.dumps(
            {
                "schema_version": "smetchik.corrections.v1",
                "source_sha256": source_hash or sha256(source),
                "corrections": [
                    {
                        "id": "COR-001",
                        "status": status,
                        "target": target or {"sheet": "ЛСР", "cell": "B2"},
                        "expected_old": expected_old,
                        "replacement": replacement,
                        "reason": reason,
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def correction_command(
    source: Path,
    corrections: Path,
    output: Path,
    changelog: Path,
) -> subprocess.CompletedProcess[str]:
    return run_cli(
        "correct",
        source,
        "--corrections-json",
        corrections,
        "--output",
        output,
        "--changelog",
        changelog,
    )


def test_confirmed_xlsx_change_creates_copy_and_changelog_without_mutating_source(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    source_before = sha256(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)
    output = tmp_path / "corrected.xlsx"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 0, completed.stderr
    assert sha256(source) == source_before
    assert load_workbook(source, data_only=False)["ЛСР"]["B2"].value == 10
    assert load_workbook(output, data_only=False)["ЛСР"]["B2"].value == 12
    log = json.loads(changelog.read_text(encoding="utf-8"))
    assert log["schema_version"] == "smetchik.changelog.v1"
    assert log["source_sha256_before"] == source_before
    assert log["source_sha256_after"] == source_before
    assert log["changes"] == [
        {
            "id": "COR-001",
            "action": "replaced",
            "target": {"sheet": "ЛСР", "cell": "B2"},
            "old_value": 10,
            "new_value": 12,
            "reason": "Подтверждено исходными документами",
        }
    ]


def test_disputed_xlsx_change_adds_comment_but_keeps_value(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    source_before = sha256(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(
        corrections,
        source,
        status="disputed",
        replacement=15,
        reason="Требуется подтверждение объёма",
    )
    output = tmp_path / "annotated.xlsx"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 0, completed.stderr
    assert sha256(source) == source_before
    cell = load_workbook(output, data_only=False)["ЛСР"]["B2"]
    assert cell.value == 10
    assert cell.comment is not None
    assert "Требуется подтверждение объёма" in cell.comment.text
    assert "15" in cell.comment.text
    log = json.loads(changelog.read_text(encoding="utf-8"))
    assert log["changes"][0]["action"] == "annotated"


def test_stale_expected_old_rejects_without_any_artifact(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    source_before = sha256(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source, expected_old=9)
    output = tmp_path / "corrected.xlsx"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 2
    assert "stale" in completed.stderr.casefold()
    assert not output.exists()
    assert not changelog.exists()
    assert sha256(source) == source_before


@pytest.mark.parametrize("existing", ["output", "changelog"])
def test_existing_artifact_is_never_overwritten(tmp_path: Path, existing: str) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)
    output = tmp_path / "corrected.xlsx"
    changelog = tmp_path / "changes.json"
    protected = output if existing == "output" else changelog
    protected.write_bytes(b"SENTINEL")

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 2
    assert "already exists" in completed.stderr.casefold()
    assert protected.read_bytes() == b"SENTINEL"
    if existing == "changelog":
        assert not output.exists()


def test_source_hash_mismatch_rejects_without_artifacts(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source, source_hash="0" * 64)
    output = tmp_path / "corrected.xlsx"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 2
    assert "hash" in completed.stderr.casefold()
    assert not output.exists()
    assert not changelog.exists()


@pytest.mark.parametrize(
    ("suffix", "payload", "target", "artifact_kind"),
    [
        (".pdf", b"%PDF-1.4\n%sidecar-only\n", {"locator": "page:1:bbox:1,2,3,4"}, "pdf_sidecar"),
        (".xml", b"<Estimate><Total>10</Total></Estimate>", {"locator": "/Estimate/Total"}, "xml_sidecar"),
        (".xls", b"legacy-xls-placeholder", {"locator": "Sheet1!B2"}, "xls_sidecar"),
    ],
)
def test_non_ooxml_formats_create_sidecar_only(
    tmp_path: Path,
    suffix: str,
    payload: bytes,
    target: dict[str, object],
    artifact_kind: str,
) -> None:
    source = tmp_path / f"source{suffix}"
    source.write_bytes(payload)
    source_before = sha256(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source, target=target)
    sidecar = tmp_path / "corrections.sidecar.json"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, sidecar, changelog)

    assert completed.returncode == 0, completed.stderr
    assert sha256(source) == source_before
    artifact = json.loads(sidecar.read_text(encoding="utf-8"))
    assert artifact["schema_version"] == "smetchik.correction-sidecar.v1"
    assert artifact["artifact_kind"] == artifact_kind
    assert artifact["source_sha256"] == source_before
    assert artifact["proposals"][0]["target"] == target
    assert artifact["proposals"][0]["applied"] is False


def test_invalid_correction_contract_is_rejected_without_traceback(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    corrections.write_text(
        '{"schema_version":"smetchik.corrections.v1","source_sha256":"'
        + sha256(source)
        + '","corrections":[{"id":"COR-001","status":"confirmed",'
        '"target":{"sheet":"ЛСР","cell":"B2"},"replacement":12,"reason":"x"}]}',
        encoding="utf-8",
    )
    output = tmp_path / "corrected.xlsx"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    combined = completed.stdout + completed.stderr
    assert completed.returncode == 2
    assert "expected_old" in combined
    assert "traceback" not in combined.casefold()
    assert not output.exists()
    assert not changelog.exists()


def test_non_finite_json_number_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    corrections.write_text(
        '{"schema_version":"smetchik.corrections.v1","source_sha256":"'
        + sha256(source)
        + '","corrections":[{"id":"COR-001","status":"confirmed",'
        '"target":{"sheet":"ЛСР","cell":"B2"},"expected_old":10,'
        '"replacement":NaN,"reason":"x"}]}',
        encoding="utf-8",
    )

    completed = correction_command(
        source,
        corrections,
        tmp_path / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "non-finite" in completed.stderr.casefold()


def test_formula_replacement_is_rejected_as_active_content(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source, replacement="=WEBSERVICE(\"https://example.invalid\")")

    completed = correction_command(
        source,
        corrections,
        tmp_path / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "formula" in completed.stderr.casefold()


def test_output_path_equal_to_source_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    before = source.read_bytes()
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)

    completed = correction_command(source, corrections, source, tmp_path / "changes.json")

    assert completed.returncode == 2
    assert "source" in completed.stderr.casefold()
    assert source.read_bytes() == before
    assert not (tmp_path / "changes.json").exists()


def test_xlsm_copy_preserves_vba_payload_without_execution_and_reports_limitation(
    tmp_path: Path,
) -> None:
    plain = tmp_path / "plain.xlsx"
    make_workbook(plain)
    source = tmp_path / "source.xlsm"
    vba_payload = b"SAFE-TEST-VBA-PAYLOAD"
    with zipfile.ZipFile(plain, "r") as input_zip, zipfile.ZipFile(source, "w") as output_zip:
        for info in input_zip.infolist():
            content = input_zip.read(info.filename)
            if info.filename == "[Content_Types].xml":
                content = content.replace(
                    b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                ).replace(
                    b"</Types>",
                    b'<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                content = content.replace(
                    b"</Relationships>",
                    b'<Relationship Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin" Id="rId999"/></Relationships>',
                )
            output_zip.writestr(info, content)
        output_zip.writestr("xl/vbaProject.bin", vba_payload)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)
    output = tmp_path / "corrected.xlsm"
    changelog = tmp_path / "changes.json"

    completed = correction_command(source, corrections, output, changelog)

    assert completed.returncode == 0, completed.stderr
    with zipfile.ZipFile(output) as archive:
        assert archive.read("xl/vbaProject.bin") == vba_payload
    log = json.loads(changelog.read_text(encoding="utf-8"))
    assert any(item["code"] == "xlsm_vba_preserved_without_execution" for item in log["limitations"])


@pytest.mark.parametrize("linked_input", ["source", "corrections"])
def test_symlinked_inputs_are_rejected(tmp_path: Path, linked_input: str) -> None:
    real_source = tmp_path / "real.xlsx"
    make_workbook(real_source)
    real_corrections = tmp_path / "real-corrections.json"
    write_corrections(real_corrections, real_source)
    source = real_source
    corrections = real_corrections
    if linked_input == "source":
        source = tmp_path / "linked.xlsx"
        source.symlink_to(real_source)
        write_corrections(real_corrections, real_source, source_hash=sha256(real_source))
    else:
        corrections = tmp_path / "linked-corrections.json"
        corrections.symlink_to(real_corrections)

    completed = correction_command(
        source,
        corrections,
        tmp_path / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "symlink" in completed.stderr.casefold()
    assert not (tmp_path / "corrected.xlsx").exists()
    assert not (tmp_path / "changes.json").exists()


def test_symlinked_output_parent_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)
    real_parent = tmp_path / "real-output"
    real_parent.mkdir()
    linked_parent = tmp_path / "linked-output"
    linked_parent.symlink_to(real_parent, target_is_directory=True)

    completed = correction_command(
        source,
        corrections,
        linked_parent / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "symlink" in completed.stderr.casefold()
    assert list(real_parent.iterdir()) == []


def test_output_and_changelog_must_be_distinct(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source)
    artifact = tmp_path / "same-path"

    completed = correction_command(source, corrections, artifact, artifact)

    assert completed.returncode == 2
    assert "distinct" in completed.stderr.casefold()
    assert not artifact.exists()


def test_corrections_json_size_is_limited_before_parsing(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    corrections.write_bytes(b"{" + b" " * (1024 * 1024 + 1) + b"}")

    completed = correction_command(
        source,
        corrections,
        tmp_path / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "size limit" in completed.stderr.casefold()


def test_source_size_is_limited_before_hashing(tmp_path: Path) -> None:
    source = tmp_path / "source.pdf"
    with source.open("wb") as stream:
        stream.truncate(100 * 1024 * 1024 + 1)
    corrections = tmp_path / "corrections.json"
    corrections.write_text("{}", encoding="utf-8")

    completed = correction_command(
        source,
        corrections,
        tmp_path / "sidecar.json",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "source size limit" in completed.stderr.casefold()


def test_correction_string_length_is_limited(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    make_workbook(source)
    corrections = tmp_path / "corrections.json"
    write_corrections(corrections, source, reason="x" * 2001)

    completed = correction_command(
        source,
        corrections,
        tmp_path / "corrected.xlsx",
        tmp_path / "changes.json",
    )

    assert completed.returncode == 2
    assert "reason" in completed.stderr.casefold()
    assert "limit" in completed.stderr.casefold()

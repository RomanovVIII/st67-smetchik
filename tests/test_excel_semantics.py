from __future__ import annotations

import json
import re
import zipfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

import excel_extract
from canonical_model import build_canonical_model
from excel_extract import extract_xls, extract_xlsx


def _rewrite_zip_member(path: Path, member_name: str, transform) -> None:
    replacement = path.with_suffix(path.suffix + ".replacement")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        replacement, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == member_name:
                payload = transform(payload)
            target.writestr(info, payload)
    replacement.replace(path)


def _set_formula_cache(path: Path, coordinate: str, cached_value: str) -> None:
    cell_pattern = re.compile(
        rb'(<c\b[^>]*\br="' + re.escape(coordinate.encode("ascii")) + rb'"[^>]*>.*?<v>)(.*?)(</v>.*?</c>)',
        re.DOTALL,
    )

    def transform(payload: bytes) -> bytes:
        changed, count = cell_pattern.subn(
            rb"\g<1>" + cached_value.encode("ascii") + rb"\g<3>", payload, count=1
        )
        assert count == 1
        return changed

    _rewrite_zip_member(path, "xl/worksheets/sheet1.xml", transform)


def test_ooxml_file_size_limit_rejects_before_openpyxl(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "large.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "x" * 200
    workbook.save(source)
    monkeypatch.setattr(excel_extract, "MAX_EXCEL_FILE_BYTES", 32)

    def must_not_load(*_args, **_kwargs):
        raise AssertionError("load_workbook must not run after preflight rejection")

    monkeypatch.setattr(excel_extract, "load_workbook", must_not_load)
    details, status, limitations, records = extract_xlsx(source)

    assert status == "rejected"
    assert records == 0
    assert details["preflight"]["file_size_bytes"] == source.stat().st_size
    limit = next(item for item in limitations if item["code"] == "excel_file_size_limit")
    assert limit["details"]["actual_bytes"] > limit["details"]["limit_bytes"]


def test_ooxml_uncompressed_and_compression_limits_reject_before_openpyxl(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "compressed.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "A" * 30_000
    workbook.save(source)
    monkeypatch.setattr(excel_extract, "MAX_OOXML_UNCOMPRESSED_BYTES", 1_000)

    def must_not_load(*_args, **_kwargs):
        raise AssertionError("load_workbook must not run after preflight rejection")

    monkeypatch.setattr(excel_extract, "load_workbook", must_not_load)
    _details, status, limitations, records = extract_xlsx(source)

    assert status == "rejected"
    assert records == 0
    assert any(item["code"] == "ooxml_uncompressed_size_limit" for item in limitations)


def test_ooxml_compression_ratio_limit_is_independent_and_precedes_openpyxl(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "ratio.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "repeated " * 1_000
    workbook.save(source)
    monkeypatch.setattr(excel_extract, "MAX_OOXML_COMPRESSION_RATIO", 1.0)

    def must_not_load(*_args, **_kwargs):
        raise AssertionError("load_workbook must not run after preflight rejection")

    monkeypatch.setattr(excel_extract, "load_workbook", must_not_load)
    details, status, limitations, records = extract_xlsx(source)

    assert status == "rejected"
    assert records == 0
    assert details["preflight"]["max_compression_ratio"] > 1.0
    assert any(item["code"] == "ooxml_compression_ratio_limit" for item in limitations)


def test_ooxml_sheet_and_cell_limits_are_checked_before_openpyxl(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "many-cells.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for index in range(1, 5):
        sheet.cell(index, 1, index)
    workbook.save(source)
    monkeypatch.setattr(excel_extract, "MAX_WORKBOOK_CELLS", 3)

    def must_not_load(*_args, **_kwargs):
        raise AssertionError("load_workbook must not run after preflight rejection")

    monkeypatch.setattr(excel_extract, "load_workbook", must_not_load)
    details, status, limitations, records = extract_xlsx(source)

    assert status == "rejected"
    assert records == 0
    assert details["preflight"]["serialized_cell_count"] == 4
    assert any(item["code"] == "excel_cell_count_limit" for item in limitations)


def test_xlsx_inventories_visibility_filters_names_tables_calculation_and_formats(
    tmp_path: Path,
) -> None:
    source = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    workbook.calculation.calcMode = "manual"
    sheet = workbook.active
    sheet.title = "ЛСР"
    sheet.append(["№", "Наименование", "Ед. изм.", "Количество", "Цена за единицу", "Сумма"])
    sheet.append([1, "Кирпич", "шт", 2, 10, "=D2*E2"])
    sheet.row_dimensions[4].hidden = True
    sheet.column_dimensions["H"].hidden = True
    sheet.auto_filter.ref = "A1:F2"
    sheet["E2"].number_format = "#,##0.00"
    table = Table(displayName="EstimateTable", ref="A1:F2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    workbook.defined_names.add(
        DefinedName("EstimateAmount", attr_text="'ЛСР'!$F$2")
    )
    workbook.save(source)

    details, status, _limitations, records = extract_xlsx(source)

    assert status in {"reliable", "partial"}
    assert details["calculation_mode"] == "manual"
    assert details["defined_names"][0]["name"] == "EstimateAmount"
    extracted_sheet = details["sheets"][0]
    assert extracted_sheet["hidden_rows"] == [4]
    assert extracted_sheet["hidden_columns"] == ["H"]
    assert extracted_sheet["auto_filter"] == "A1:F2"
    assert extracted_sheet["tables"][0]["name"] == "EstimateTable"
    assert extracted_sheet["tables"][0]["ref"] == "A1:F2"
    assert extracted_sheet["number_formats"]["#,##0.00"] == ["E2"]
    assert records == 2  # one semantic row plus one formula


def test_xlsx_emits_calculation_basis_only_for_an_explicit_recognized_row_formula(
    tmp_path: Path,
) -> None:
    source = tmp_path / "explicit-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ЛСР"
    sheet.append(["Код", "Наименование", "Ед. изм.", "Количество", "Цена за единицу", "Сумма"])
    sheet.append(["R-1", "Работа", "м2", 2, 10, "=D2*E2"])
    sheet.append(["R-2", "Работа без формулы", "м2", 3, 10, 30])
    workbook.save(source)

    details, _status, _limitations, _records = extract_xlsx(source)

    rows = details["sheets"][0]["semantic_tables"][0]["rows"]
    assert rows[0]["calculation_basis"] == {
        "formula": "quantity * unit_price",
        "operand_fields": ["quantity", "unit_price"],
        "source_fields_verified_complete": True,
        "evidence": {
            "sheet": "ЛСР",
            "cell_range": "F2",
            "locator": "ЛСР!F2",
        },
    }
    assert "calculation_basis" not in rows[1]


def test_safe_formula_evaluator_supports_dependencies_ranges_and_required_functions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "formulas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["B1"] = "=SUM(A1:A2)"
    sheet["B2"] = "=ROUND(B1/2,0)"
    sheet["C1"] = "=MIN(A1:A2,B1)"
    sheet["C2"] = "=MAX(A1,A2,B2)"
    sheet["D1"] = "=B2+C1+C2"
    sheet["D2"] = "=Calc!A1"
    sheet["E1"] = "=E2"
    sheet["E2"] = "=E1"
    workbook.save(source)

    details, status, limitations, records = extract_xlsx(source)

    cells = {cell["coordinate"]: cell for cell in details["sheets"][0]["cells"]}
    assert cells["B1"]["computed_safe_value"] == 3
    assert cells["B2"]["computed_safe_value"] == 2
    assert cells["C1"]["computed_safe_value"] == 1
    assert cells["C2"]["computed_safe_value"] == 2
    assert cells["D1"]["computed_safe_value"] == 5
    assert cells["D2"]["computed_safe_value"] is None
    assert cells["E1"]["computed_safe_value"] is None
    assert status == "partial"
    assert records == 8
    codes = [item["code"] for item in limitations]
    assert "unsupported_formula" in codes
    assert "formula_cycle" in codes


def test_formula_cache_mismatch_is_reported_as_raw_detail_not_finding(
    tmp_path: Path,
) -> None:
    source = tmp_path / "cache.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = 10
    sheet["B1"] = 5
    sheet["C1"] = "=A1+B1"
    workbook.save(source)
    _set_formula_cache(source, "C1", "14")

    details, _status, _limitations, records = extract_xlsx(source)

    mismatch = details["formula_cache_mismatches"][0]
    assert mismatch == {
        "locator": "Calc!C1",
        "formula": "=A1+B1",
        "cached_value": 14,
        "computed_safe_value": 15,
        "raw_difference": 1,
    }
    cell = next(
        cell for cell in details["sheets"][0]["cells"] if cell["coordinate"] == "C1"
    )
    assert cell["cached_value"] == 14
    assert cell["computed_safe_value"] == 15
    assert records == 1


def test_decimal_formula_equal_to_cache_does_not_create_float_noise_mismatch(
    tmp_path: Path,
) -> None:
    source = tmp_path / "decimal-cache.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = 0.1
    sheet["B1"] = 0.2
    sheet["C1"] = "=A1+B1"
    workbook.save(source)
    _set_formula_cache(source, "C1", "0.3")

    details, _status, _limitations, records = extract_xlsx(source)

    cell = next(
        cell for cell in details["sheets"][0]["cells"] if cell["coordinate"] == "C1"
    )
    assert cell["cached_value"] == 0.3
    assert cell["computed_safe_value"] == 0.3
    assert details["formula_cache_mismatches"] == []
    assert records == 1


def test_decimal_formula_real_cache_mismatch_keeps_exact_normalized_difference(
    tmp_path: Path,
) -> None:
    source = tmp_path / "decimal-real-mismatch.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = 0.1
    sheet["B1"] = 0.2
    sheet["C1"] = "=A1+B1"
    workbook.save(source)
    _set_formula_cache(source, "C1", "0.4")

    details, _status, _limitations, _records = extract_xlsx(source)

    assert details["formula_cache_mismatches"] == [
        {
            "locator": "Calc!C1",
            "formula": "=A1+B1",
            "cached_value": 0.4,
            "computed_safe_value": 0.3,
            "raw_difference": -0.1,
        }
    ]


def test_semantic_rows_require_explicit_recognized_headers_and_keep_locators(
    tmp_path: Path,
) -> None:
    source = tmp_path / "semantic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estimate"
    sheet.append(
        [
            "№",
            "Код",
            "Наименование",
            "Ед. изм.",
            "Количество",
            "Цена за единицу",
            "Сумма",
            "Индекс",
            "Коэффициент",
            "Категория",
            "Источник",
            "Дата",
            "НДС",
            "Доставка",
            "Налоговая база",
        ]
    )
    sheet.append([1, "ГЭСН", "Работа", "м3", 2, 10, 20, 1.1, 1.2, "work", "ФГИС", "2026-08-26", 4.4, 1, 20])
    sheet.append([2, "ФЕР", "Материал", "шт", 3, 5, 15, None, None, "material", "price", None, 3.3, 2, 15])
    unknown = workbook.create_sheet("Unknown")
    unknown.append(["Alpha", "Beta"])
    unknown.append([1, 2])
    workbook.save(source)

    details, _status, _limitations, records = extract_xlsx(source)

    estimate = details["sheets"][0]
    assert estimate["semantic_tables"][0]["header_row"] == 1
    assert set(estimate["semantic_tables"][0]["columns"]) == {
        "number",
        "code",
        "name",
        "unit",
        "quantity",
        "unit_price",
        "amount",
        "index",
        "coefficient",
        "category",
        "source",
        "date",
        "vat",
        "delivery",
        "taxable_base",
    }
    first = estimate["semantic_tables"][0]["rows"][0]
    assert first["field_values"]["quantity"] == 2
    assert first["cell_locators"]["quantity"] == "Estimate!E2"
    assert details["sheets"][1]["semantic_tables"] == []
    assert records == 2


def test_generic_number_name_registry_is_not_an_estimate_table(tmp_path: Path) -> None:
    source = tmp_path / "registry.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet.append(["Номер", "Наименование"])
    sheet.append([1, "Паспорт объекта"])
    sheet.append([2, "Сопроводительное письмо"])
    workbook.save(source)

    details, _status, _limitations, records = extract_xlsx(source)

    assert details["semantic_row_count"] == 0
    assert details["sheets"][0]["semantic_tables"] == []
    assert records == 0
    canonical = build_canonical_model(
        [{"path": source.name, "file_type": "xlsx", "details": details}],
        {},
    )
    assert canonical["rows"] == []
    assert canonical["unclassified_candidate_ranges"] == [
        {
            "source_path": source.name,
            "sheet": "Реестр",
            "locator": "Реестр!used-range",
            "reason": "spreadsheet_content_not_classified_as_estimate_rows",
        }
    ]


def test_quantity_takeoff_signature_remains_semantic_without_cost_columns(
    tmp_path: Path,
) -> None:
    source = tmp_path / "quantity-takeoff.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ВОР"
    sheet.append(["№", "Наименование", "Ед. изм.", "Количество"])
    sheet.append([1, "Разборка покрытия", "м2", 12.5])
    workbook.save(source)

    details, _status, _limitations, records = extract_xlsx(source)

    table = details["sheets"][0]["semantic_tables"][0]
    assert set(table["columns"]) == {"number", "name", "unit", "quantity"}
    assert table["rows"][0]["field_values"]["quantity"] == 12.5
    assert records == 1


def test_xls_file_size_limit_rejects_before_xlrd(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"x" * 100)
    monkeypatch.setattr(excel_extract, "MAX_EXCEL_FILE_BYTES", 10)

    def must_not_open(*_args, **_kwargs):
        raise AssertionError("xlrd must not run after preflight rejection")

    monkeypatch.setattr(excel_extract.xlrd, "open_workbook", must_not_open)
    details, status, limitations, records = extract_xls(source)

    assert status == "rejected"
    assert records == 0
    assert details["read_only"] is True
    assert limitations[0]["code"] == "excel_file_size_limit"


def test_xls_inventories_available_hidden_metadata_names_and_semantic_rows(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"bounded fake xls")

    class FakeSheet:
        name = "ЛСР"
        visibility = 1
        nrows = 2
        ncols = 3
        rowinfo_map = {1: SimpleNamespace(hidden=True)}
        colinfo_map = {2: SimpleNamespace(hidden=True)}
        values = [["№", "Name", "Amount"], [1, "Работа", 100]]

        def cell_value(self, row: int, column: int):
            return self.values[row][column]

    class FakeWorkbook:
        nsheets = 1
        name_obj_list = [
            SimpleNamespace(
                name="EstimateTotal", formula_text="ЛСР!$C$2", scope=-1, hidden=False
            )
        ]
        xf_list: list = []
        format_map: dict = {}
        released = False

        def sheets(self):
            return [FakeSheet()]

        def release_resources(self):
            self.released = True

    fake = FakeWorkbook()

    def open_read_only(path, *, on_demand, formatting_info):
        assert path == source
        assert on_demand is True
        assert formatting_info is True
        return fake

    monkeypatch.setattr(excel_extract.xlrd, "open_workbook", open_read_only)
    details, status, limitations, records = extract_xls(source)

    sheet = details["sheets"][0]
    assert status == "partial"
    assert limitations[0]["code"] == "legacy_xls_formula_metadata_unavailable"
    assert sheet["state"] == "hidden"
    assert sheet["hidden_rows"] == [2]
    assert sheet["hidden_columns"] == [3]
    assert sheet["semantic_tables"][0]["rows"][0]["cell_locators"]["amount"] == "ЛСР!R2C3"
    assert details["defined_names"][0]["name"] == "EstimateTotal"
    assert records == 1
    assert fake.released is True


def test_xls_declared_cell_limit_prevents_cell_iteration(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"bounded fake xls")

    class FakeSheet:
        name = "Huge"
        nrows = 100
        ncols = 100

        def cell_value(self, _row: int, _column: int):
            raise AssertionError("cells must not be iterated after dimension rejection")

    class FakeWorkbook:
        nsheets = 1
        released = False

        def sheets(self):
            return [FakeSheet()]

        def release_resources(self):
            self.released = True

    fake = FakeWorkbook()
    monkeypatch.setattr(excel_extract, "MAX_WORKBOOK_CELLS", 10)
    monkeypatch.setattr(excel_extract.xlrd, "open_workbook", lambda *_args, **_kwargs: fake)

    details, status, limitations, records = extract_xls(source)

    assert status == "rejected"
    assert records == 0
    assert details["declared_cell_count"] == 10_000
    assert limitations[0]["code"] == "excel_cell_count_limit"
    assert fake.released is True


def test_sheet_limit_is_enforced_before_cell_iteration(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "sheets.xlsx"
    workbook = Workbook()
    workbook.create_sheet("Second")
    workbook.save(source)
    monkeypatch.setattr(excel_extract, "MAX_WORKBOOK_SHEETS", 1)

    details, status, limitations, records = extract_xlsx(source)

    assert status == "rejected"
    assert records == 0
    assert details["preflight"]["worksheet_count"] == 2
    assert any(item["code"] == "excel_sheet_count_limit" for item in limitations)


def test_declared_dimension_limit_catches_sparse_malicious_sheet(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "dimension.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = 1
    workbook.save(source)
    _rewrite_zip_member(
        source,
        "xl/worksheets/sheet1.xml",
        lambda payload: re.sub(
            rb"(<dimension\b[^>]*\bref=)[\"'][^\"']+[\"']",
            rb'\g<1>"A1:Z100"',
            payload,
            count=1,
        ),
    )
    monkeypatch.setattr(excel_extract, "MAX_WORKBOOK_DECLARED_CELLS", 1_000)

    details, status, limitations, _records = extract_xlsx(source)

    assert status == "rejected"
    assert details["preflight"]["largest_declared_range_cells"] == 2_600
    assert any(item["code"] == "excel_declared_dimension_limit" for item in limitations)


def test_formula_record_count_does_not_count_plain_unknown_cells(tmp_path: Path) -> None:
    source = tmp_path / "counts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "free text"
    sheet["B2"] = 3
    sheet["C2"] = "=B2*2"
    workbook.save(source)

    _details, _status, _limitations, records = extract_xlsx(source)

    assert records == 1


def test_excel_dates_are_emitted_as_json_scalars(tmp_path: Path) -> None:
    source = tmp_path / "dates.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = date(2026, 8, 26)
    workbook.save(source)

    details, _status, _limitations, records = extract_xlsx(source)

    assert details["sheets"][0]["cells"][0]["value"] == "2026-08-26T00:00:00"
    json.dumps(details, allow_nan=False)
    assert records == 0

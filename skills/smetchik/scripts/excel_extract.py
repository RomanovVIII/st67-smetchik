from __future__ import annotations

import math
import re
import zipfile
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Sequence

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


MAX_EXCEL_FILE_BYTES = 100 * 1024 * 1024
MAX_OOXML_ENTRIES = 10_000
MAX_OOXML_UNCOMPRESSED_BYTES = 500 * 1024 * 1024
MAX_OOXML_SINGLE_MEMBER_BYTES = 100 * 1024 * 1024
MAX_OOXML_COMPRESSION_RATIO = 200.0
MAX_WORKBOOK_SHEETS = 512
MAX_WORKBOOK_CELLS = 5_000_000
MAX_WORKBOOK_DECLARED_CELLS = 20_000_000
MAX_FORMULA_RANGE_CELLS = 100_000

CELL_REFERENCE = re.compile(r"^\$?([A-Z]{1,3})\$?([1-9][0-9]*)$", re.IGNORECASE)
CELL_TAG = re.compile(rb"<(?:[A-Za-z_][\w.-]*:)?c(?:\s|>)")
DIMENSION_REF = re.compile(
    rb"<(?:[A-Za-z_][\w.-]*:)?dimension\b[^>]*\bref=[\"']([^\"']+)[\"']"
)
TOKEN = re.compile(
    r"\s*(?:(?P<number>(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][+-]?\d+)?)"
    r"|(?P<cell>\$?[A-Za-z]{1,3}\$?[1-9][0-9]*)"
    r"|(?P<identifier>[A-Za-z_][A-Za-z0-9_.]*)"
    r"|(?P<operator>[+\-*/(),:]))"
)

HEADER_ALIASES: dict[str, set[str]] = {
    "number": {"№", "№ п/п", "n", "no", "number", "номер", "п/п"},
    "code": {"код", "обоснование", "шифр", "code", "rate code"},
    "name": {
        "наименование",
        "название",
        "описание",
        "name",
        "description",
        "work description",
    },
    "unit": {"ед. изм.", "ед изм", "единица измерения", "unit", "uom"},
    "quantity": {"количество", "объем", "qty", "quantity", "volume"},
    "unit_price": {
        "цена за единицу",
        "единичная цена",
        "цена ед.",
        "unit price",
        "unit_price",
        "price per unit",
    },
    "amount": {"стоимость", "сумма", "всего", "amount", "total", "cost"},
    "index": {"индекс", "index"},
    "coefficient": {"коэффициент", "коэф.", "коэф", "coeff", "coefficient"},
    "category": {"категория", "category"},
    "source": {"источник", "source"},
    "date": {"дата", "date"},
    "vat": {"ндс", "vat"},
    "delivery": {"доставка", "delivery"},
    "taxable_base": {
        "налоговая база",
        "база ндс",
        "облагаемая база",
        "taxable base",
        "taxable_base",
    },
}
HEADER_LOOKUP = {
    alias: field for field, aliases in HEADER_ALIASES.items() for alias in aliases
}
SEMANTIC_ANCHORS = {"quantity", "unit_price", "amount"}
CALCULATION_FORMULAS = {
    ("quantity", "unit_price"): "quantity * unit_price",
    ("quantity", "unit_price", "index"): "quantity * unit_price * index",
    ("quantity", "unit_price", "coefficient"): "quantity * unit_price * coefficient",
    (
        "quantity",
        "unit_price",
        "index",
        "coefficient",
    ): "quantity * unit_price * index * coefficient",
}


class UnsupportedFormula(ValueError):
    pass


class FormulaCycle(UnsupportedFormula):
    pass


class _RangeValue(list[Decimal]):
    pass


def _json_scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _decimal_number(value: Decimal | float | int) -> Decimal:
    if isinstance(value, bool) or not isinstance(value, (Decimal, int, float)):
        raise UnsupportedFormula("non-numeric result")
    try:
        decimal_value = value if isinstance(value, Decimal) else Decimal(str(value))
    except InvalidOperation as error:
        raise UnsupportedFormula("invalid number") from error
    if not decimal_value.is_finite():
        raise UnsupportedFormula("non-finite result")
    return decimal_value


def _normalise_number(value: Decimal | float | int) -> int | float:
    decimal_value = _decimal_number(value)
    if decimal_value == decimal_value.to_integral_value():
        return int(decimal_value)
    normalized = float(decimal_value)
    if not math.isfinite(normalized):
        raise UnsupportedFormula("result exceeds the finite JSON number range")
    return normalized


def _normalise_reference(reference: str) -> str:
    match = CELL_REFERENCE.fullmatch(reference)
    if match is None:
        raise UnsupportedFormula(f"invalid cell reference: {reference}")
    return f"{match.group(1).upper()}{match.group(2)}"


def _tokenise_formula(formula: str) -> list[tuple[str, str]]:
    expression = formula.removeprefix("=").strip()
    if not expression:
        raise UnsupportedFormula("empty formula")
    if any(marker in expression for marker in ("!", "[", "]", '"', "'", ";")):
        raise UnsupportedFormula("external, cross-sheet or textual reference")
    tokens: list[tuple[str, str]] = []
    offset = 0
    while offset < len(expression):
        match = TOKEN.match(expression, offset)
        if match is None:
            raise UnsupportedFormula(f"unsupported token at offset {offset}")
        kind = match.lastgroup
        assert kind is not None
        tokens.append((kind, match.group(kind)))
        offset = match.end()
    return tokens


class _FormulaParser:
    def __init__(
        self,
        tokens: Sequence[tuple[str, str]],
        resolve_cell: Callable[[str], Decimal],
        resolve_range: Callable[[str, str], list[Decimal]],
    ) -> None:
        self.tokens = tokens
        self.resolve_cell = resolve_cell
        self.resolve_range = resolve_range
        self.position = 0

    def _peek(self, kind: str | None = None, value: str | None = None) -> bool:
        if self.position >= len(self.tokens):
            return False
        token_kind, token_value = self.tokens[self.position]
        return (kind is None or token_kind == kind) and (
            value is None or token_value.upper() == value.upper()
        )

    def _take(self, kind: str | None = None, value: str | None = None) -> tuple[str, str]:
        if not self._peek(kind, value):
            raise UnsupportedFormula(f"expected {value or kind or 'token'}")
        token = self.tokens[self.position]
        self.position += 1
        return token

    def parse(self) -> Decimal:
        value = self._expression()
        if self.position != len(self.tokens):
            raise UnsupportedFormula("trailing expression")
        return _decimal_number(value)

    def _expression(self) -> Decimal:
        value = self._term()
        while self._peek("operator", "+") or self._peek("operator", "-"):
            operator = self._take("operator")[1]
            right = self._term()
            value = value + right if operator == "+" else value - right
        return value

    def _term(self) -> Decimal:
        value = self._unary()
        while self._peek("operator", "*") or self._peek("operator", "/"):
            operator = self._take("operator")[1]
            right = self._unary()
            if operator == "/" and right == 0:
                raise UnsupportedFormula("division by zero")
            value = value * right if operator == "*" else value / right
        return value

    def _unary(self) -> Decimal:
        if self._peek("operator", "+"):
            self._take("operator", "+")
            return self._unary()
        if self._peek("operator", "-"):
            self._take("operator", "-")
            return -self._unary()
        return self._primary()

    def _primary(self) -> Decimal:
        if self._peek("number"):
            raw = self._take("number")[1]
            try:
                return _decimal_number(Decimal(raw))
            except InvalidOperation as error:
                raise UnsupportedFormula("invalid number") from error
        if self._peek("cell"):
            reference = _normalise_reference(self._take("cell")[1])
            if self._peek("operator", ":"):
                raise UnsupportedFormula("range is valid only as a function argument")
            return self.resolve_cell(reference)
        if self._peek("identifier"):
            return self._function()
        if self._peek("operator", "("):
            self._take("operator", "(")
            value = self._expression()
            self._take("operator", ")")
            return value
        raise UnsupportedFormula("expected a numeric literal, cell or supported function")

    def _argument(self) -> Decimal | _RangeValue:
        if (
            self._peek("cell")
            and self.position + 1 < len(self.tokens)
            and self.tokens[self.position + 1] == ("operator", ":")
        ):
            start = _normalise_reference(self._take("cell")[1])
            self._take("operator", ":")
            end = _normalise_reference(self._take("cell")[1])
            return _RangeValue(self.resolve_range(start, end))
        return self._expression()

    def _function(self) -> Decimal:
        name = self._take("identifier")[1].upper()
        if name not in {"SUM", "ROUND", "MIN", "MAX"}:
            raise UnsupportedFormula(f"unsupported function: {name}")
        self._take("operator", "(")
        arguments: list[Decimal | _RangeValue] = []
        if not self._peek("operator", ")"):
            arguments.append(self._argument())
            while self._peek("operator", ","):
                self._take("operator", ",")
                arguments.append(self._argument())
        self._take("operator", ")")
        flattened: list[Decimal] = []
        for argument in arguments:
            flattened.extend(argument if isinstance(argument, _RangeValue) else [argument])
        if name == "SUM":
            return sum(flattened, Decimal(0))
        if name in {"MIN", "MAX"}:
            if not flattened:
                raise UnsupportedFormula(f"{name} requires at least one argument")
            return min(flattened) if name == "MIN" else max(flattened)
        if len(arguments) != 2 or any(isinstance(item, _RangeValue) for item in arguments):
            raise UnsupportedFormula("ROUND requires two scalar arguments")
        value, digits = arguments
        if int(digits) != digits or abs(int(digits)) > 15:
            raise UnsupportedFormula("ROUND digits must be an integer from -15 to 15")
        try:
            rounded = value.quantize(
                Decimal("1").scaleb(-int(digits)), rounding=ROUND_HALF_UP
            )
        except InvalidOperation as error:
            raise UnsupportedFormula("ROUND result is not representable") from error
        return rounded


def _evaluate_formula(
    formula: str,
    resolve_cell: Callable[[str], Decimal],
    resolve_range: Callable[[str, str], list[Decimal]],
) -> Decimal:
    return _FormulaParser(_tokenise_formula(formula), resolve_cell, resolve_range).parse()


def evaluate_safe_formula(formula: str, values: dict[str, Any]) -> int | float:
    normalised = {_normalise_reference(key): value for key, value in values.items()}

    def resolve_cell(reference: str) -> Decimal:
        value = normalised.get(reference)
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise UnsupportedFormula(f"non-numeric or missing dependency: {reference}")
        return _decimal_number(value)

    def resolve_range(start: str, end: str) -> list[Decimal]:
        min_column, min_row, max_column, max_row = range_boundaries(f"{start}:{end}")
        count = (max_column - min_column + 1) * (max_row - min_row + 1)
        if count > MAX_FORMULA_RANGE_CELLS:
            raise UnsupportedFormula("formula range exceeds the safety limit")
        return [
            resolve_cell(f"{get_column_letter(column)}{row}")
            for row in range(min_row, max_row + 1)
            for column in range(min_column, max_column + 1)
        ]

    return _normalise_number(_evaluate_formula(formula, resolve_cell, resolve_range))


def _limit(
    code: str,
    message: str,
    locator: str,
    **details: Any,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "code": code,
        "message": message,
        "evidence": [{"locator": locator}],
    }
    if details:
        item["details"] = details
    return item


def _declared_range_cells(reference: str) -> int:
    min_column, min_row, max_column, max_row = range_boundaries(reference)
    return (max_column - min_column + 1) * (max_row - min_row + 1)


def _normalise_header(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.casefold().replace("ё", "е").replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text).strip(":")
    return text or None


def _header_mapping(values: dict[int, Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column, value in values.items():
        field = HEADER_LOOKUP.get(_normalise_header(value) or "")
        if field is not None and field not in mapping:
            mapping[field] = column
    if len(mapping) < 3 or not (set(mapping) & SEMANTIC_ANCHORS):
        return {}
    return mapping


def _recognized_row_calculation_basis(
    *,
    sheet_name: str,
    row_number: int,
    mapping: dict[str, int],
    formulas: dict[str, str],
    locators: Callable[[int, int], str],
) -> dict[str, Any] | None:
    amount_column = mapping.get("amount")
    if amount_column is None:
        return None
    amount_coordinate = f"{get_column_letter(amount_column)}{row_number}".upper()
    formula = formulas.get(amount_coordinate)
    if not isinstance(formula, str):
        return None
    expression = formula.removeprefix("=").replace("$", "")
    factors = [item.strip().upper() for item in expression.split("*")]
    if not factors or any(CELL_REFERENCE.fullmatch(item) is None for item in factors):
        return None
    field_by_coordinate = {
        f"{get_column_letter(column)}{row_number}".upper(): field
        for field, column in mapping.items()
        if field in {"quantity", "unit_price", "index", "coefficient"}
    }
    if len(factors) != len(set(factors)) or set(factors) - field_by_coordinate.keys():
        return None
    fields_present = {field_by_coordinate[factor] for factor in factors}
    operands = tuple(
        field
        for field in ("quantity", "unit_price", "index", "coefficient")
        if field in fields_present
    )
    canonical_formula = CALCULATION_FORMULAS.get(operands)
    if canonical_formula is None or len(operands) != len(factors):
        return None
    return {
        "formula": canonical_formula,
        "operand_fields": list(operands),
        "source_fields_verified_complete": True,
        "evidence": {
            "sheet": sheet_name,
            "cell_range": amount_coordinate,
            "locator": locators(row_number, amount_column),
        },
    }


def _semantic_tables(
    *,
    sheet_name: str,
    occupied: dict[tuple[int, int], Any],
    locators: Callable[[int, int], str],
    explicit_tables: list[dict[str, Any]],
    formulas: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    formulas = formulas or {}
    rows: dict[int, dict[int, Any]] = {}
    for (row, column), value in occupied.items():
        rows.setdefault(row, {})[column] = value
    explicit_by_header: dict[int, tuple[int, str | None]] = {}
    for table in explicit_tables:
        try:
            min_column, min_row, max_column, max_row = range_boundaries(table["ref"])
        except (KeyError, TypeError, ValueError):
            continue
        if (max_column - min_column + 1) * (max_row - min_row + 1) <= MAX_WORKBOOK_CELLS:
            explicit_by_header[min_row] = (max_row, table.get("name"))
    detected: list[dict[str, Any]] = []
    occupied_rows = sorted(rows)
    for header_row in occupied_rows:
        mapping = _header_mapping(rows[header_row])
        if not mapping:
            continue
        explicit = explicit_by_header.get(header_row)
        if explicit is not None:
            end_row, table_name = explicit
            candidates: Sequence[int] = range(header_row + 1, end_row + 1)
        else:
            table_name = None
            consecutive: list[int] = []
            expected = header_row + 1
            for row_number in occupied_rows:
                if row_number <= header_row:
                    continue
                if row_number != expected or _header_mapping(rows[row_number]):
                    break
                consecutive.append(row_number)
                expected += 1
            candidates = consecutive
        semantic_rows: list[dict[str, Any]] = []
        for row_number in candidates:
            field_values = {
                field: occupied.get((row_number, column)) for field, column in mapping.items()
            }
            if not any(value not in (None, "") for value in field_values.values()):
                if explicit is None:
                    break
                continue
            semantic_row = {
                "row": row_number,
                "field_values": field_values,
                "cell_locators": {
                    field: locators(row_number, column) for field, column in mapping.items()
                },
                "evidence": {"locator": f"{sheet_name}!{row_number}:{row_number}"},
            }
            calculation_basis = _recognized_row_calculation_basis(
                sheet_name=sheet_name,
                row_number=row_number,
                mapping=mapping,
                formulas=formulas,
                locators=locators,
            )
            if calculation_basis is not None:
                semantic_row["calculation_basis"] = calculation_basis
            semantic_rows.append(semantic_row)
        if semantic_rows:
            detected.append(
                {
                    "name": table_name,
                    "header_row": header_row,
                    "columns": {
                        field: {
                            "column": column,
                            "header": rows[header_row][column],
                            "locator": locators(header_row, column),
                        }
                        for field, column in mapping.items()
                    },
                    "rows": semantic_rows,
                }
            )
    return detected


def _defined_names(workbook: Any) -> list[dict[str, Any]]:
    return [
        {
            "name": item.name,
            "target": item.attr_text,
            "local_sheet_id": item.localSheetId,
            "hidden": bool(item.hidden),
        }
        for item in workbook.defined_names.values()
    ]


def _table_inventory(sheet: Any) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    for table in sheet.tables.values():
        style = table.tableStyleInfo
        tables.append(
            {
                "name": table.displayName or table.name,
                "ref": table.ref,
                "style": style.name if style is not None else None,
                "show_row_stripes": bool(style.showRowStripes) if style is not None else False,
            }
        )
    return tables


def _compute_sheet_formulas(
    formulas: dict[str, str], scalar_values: dict[str, Any]
) -> tuple[dict[str, int | float], dict[str, UnsupportedFormula]]:
    computed: dict[str, Decimal] = {}
    errors: dict[str, UnsupportedFormula] = {}
    active: list[str] = []

    def resolve_cell(reference: str) -> Decimal:
        reference = _normalise_reference(reference)
        if reference in computed:
            return computed[reference]
        if reference in active:
            cycle = active[active.index(reference) :] + [reference]
            raise FormulaCycle("dependency cycle: " + " -> ".join(cycle))
        scalar = scalar_values.get(reference)
        if isinstance(scalar, bool) or not isinstance(scalar, (int, float)):
            if reference not in formulas:
                raise UnsupportedFormula(f"non-numeric or missing dependency: {reference}")
            active.append(reference)
            try:
                value = _evaluate_formula(formulas[reference], resolve_cell, resolve_range)
                computed[reference] = value
                return value
            finally:
                active.pop()
        return _decimal_number(scalar)

    def resolve_range(start: str, end: str) -> list[Decimal]:
        min_column, min_row, max_column, max_row = range_boundaries(f"{start}:{end}")
        count = (max_column - min_column + 1) * (max_row - min_row + 1)
        if count > MAX_FORMULA_RANGE_CELLS:
            raise UnsupportedFormula("formula range exceeds the safety limit")
        return [
            resolve_cell(f"{get_column_letter(column)}{row}")
            for row in range(min_row, max_row + 1)
            for column in range(min_column, max_column + 1)
        ]

    for coordinate in sorted(formulas):
        try:
            resolve_cell(coordinate)
        except UnsupportedFormula as error:
            errors[coordinate] = error
    return {
        coordinate: _normalise_number(value) for coordinate, value in computed.items()
    }, errors


def _base_openxml_details(
    path: Path, preflight: dict[str, Any], macro_payload: bool
) -> dict[str, Any]:
    return {
        "format": path.suffix.lower().lstrip("."),
        "read_only": True,
        "preflight": preflight,
        "sheets": [],
        "hidden_sheets": [],
        "merged_cells": [],
        "defined_names": [],
        "calculation_mode": None,
        "macro_capable": path.suffix.lower() == ".xlsm",
        "vba_payload_present": macro_payload,
        "macros_executed": False,
        "external_links_detected": False,
        "missing_cached_formula_values": [],
        "formula_cache_mismatches": [],
    }


def _ooxml_preflight(
    path: Path,
) -> tuple[dict[str, Any], list[dict[str, Any]], bool, bool]:
    size = path.stat().st_size
    details: dict[str, Any] = {
        "file_size_bytes": size,
        "archive_entry_count": 0,
        "declared_uncompressed_bytes": 0,
        "max_compression_ratio": 0.0,
        "zero_compressed_member": False,
        "worksheet_count": 0,
        "serialized_cell_count": 0,
        "largest_declared_range_cells": 0,
    }
    if size > MAX_EXCEL_FILE_BYTES:
        return details, [
            _limit(
                "excel_file_size_limit",
                "Размер книги превышает защитный лимит; книга не открывалась.",
                path.name,
                actual_bytes=size,
                limit_bytes=MAX_EXCEL_FILE_BYTES,
            )
        ], False, False
    try:
        with zipfile.ZipFile(path) as archive:
            members = [item for item in archive.infolist() if not item.is_dir()]
            uncompressed = sum(item.file_size for item in members)
            zero_compressed_member = any(
                item.file_size > 0 and item.compress_size == 0 for item in members
            )
            ratios = [
                item.file_size / item.compress_size
                for item in members
                if item.file_size and item.compress_size
            ]
            worksheets = [
                item
                for item in members
                if re.fullmatch(r"xl/worksheets/[^/]+\.xml", item.filename, re.IGNORECASE)
            ]
            details.update(
                archive_entry_count=len(members),
                declared_uncompressed_bytes=uncompressed,
                max_compression_ratio=max(ratios, default=0.0),
                zero_compressed_member=zero_compressed_member,
                worksheet_count=len(worksheets),
            )
            macro = any(item.filename.casefold().endswith("vbaproject.bin") for item in members)
            external = any("/externallinks/" in item.filename.casefold() for item in members)
            limits: list[dict[str, Any]] = []
            limit_checks = (
                (
                    len(members) > MAX_OOXML_ENTRIES,
                    "ooxml_entry_count_limit",
                    "Число частей OOXML превышает защитный лимит; книга не открывалась.",
                    {"actual_entries": len(members), "limit_entries": MAX_OOXML_ENTRIES},
                ),
                (
                    uncompressed > MAX_OOXML_UNCOMPRESSED_BYTES,
                    "ooxml_uncompressed_size_limit",
                    "Заявленный распакованный размер OOXML превышает защитный лимит; книга не открывалась.",
                    {
                        "actual_bytes": uncompressed,
                        "limit_bytes": MAX_OOXML_UNCOMPRESSED_BYTES,
                    },
                ),
                (
                    max((item.file_size for item in members), default=0)
                    > MAX_OOXML_SINGLE_MEMBER_BYTES,
                    "ooxml_member_size_limit",
                    "Одна часть OOXML превышает защитный лимит; книга не открывалась.",
                    {
                        "actual_bytes": max((item.file_size for item in members), default=0),
                        "limit_bytes": MAX_OOXML_SINGLE_MEMBER_BYTES,
                    },
                ),
                (
                    zero_compressed_member
                    or details["max_compression_ratio"] > MAX_OOXML_COMPRESSION_RATIO,
                    "ooxml_compression_ratio_limit",
                    "Коэффициент сжатия OOXML превышает защитный лимит; книга не открывалась.",
                    {
                        "actual_ratio": (
                            "infinite"
                            if zero_compressed_member
                            else details["max_compression_ratio"]
                        ),
                        "limit_ratio": MAX_OOXML_COMPRESSION_RATIO,
                    },
                ),
                (
                    len(worksheets) > MAX_WORKBOOK_SHEETS,
                    "excel_sheet_count_limit",
                    "Число листов превышает защитный лимит; книга не открывалась.",
                    {
                        "actual_sheets": len(worksheets),
                        "limit_sheets": MAX_WORKBOOK_SHEETS,
                    },
                ),
            )
            for exceeded, code, message, values in limit_checks:
                if exceeded:
                    limits.append(_limit(code, message, path.name, **values))
            if limits:
                return details, limits, macro, external
            for member in worksheets:
                payload = archive.read(member)
                details["serialized_cell_count"] += len(CELL_TAG.findall(payload))
                dimension = DIMENSION_REF.search(payload)
                if dimension is not None:
                    try:
                        declared = _declared_range_cells(dimension.group(1).decode("ascii"))
                    except (UnicodeDecodeError, ValueError):
                        declared = MAX_WORKBOOK_DECLARED_CELLS + 1
                    details["largest_declared_range_cells"] = max(
                        details["largest_declared_range_cells"], declared
                    )
            if details["serialized_cell_count"] > MAX_WORKBOOK_CELLS:
                limits.append(
                    _limit(
                        "excel_cell_count_limit",
                        "Число сериализованных ячеек превышает защитный лимит; книга не открывалась.",
                        path.name,
                        actual_cells=details["serialized_cell_count"],
                        limit_cells=MAX_WORKBOOK_CELLS,
                    )
                )
            if details["largest_declared_range_cells"] > MAX_WORKBOOK_DECLARED_CELLS:
                limits.append(
                    _limit(
                        "excel_declared_dimension_limit",
                        "Объявленный диапазон листа превышает защитный лимит; книга не открывалась.",
                        path.name,
                        actual_cells=details["largest_declared_range_cells"],
                        limit_cells=MAX_WORKBOOK_DECLARED_CELLS,
                    )
                )
            return details, limits, macro, external
    except (OSError, zipfile.BadZipFile, RuntimeError):
        return details, [
            _limit(
                "unreadable_openxml_archive",
                "Контейнер OOXML повреждён или не является безопасно читаемым ZIP.",
                path.name,
            )
        ], False, False


def extract_xlsx(path: Path) -> tuple[dict[str, Any], str, list[dict[str, Any]], int]:
    preflight, preflight_limits, macro_payload, archive_external_links = (
        _ooxml_preflight(path)
    )
    details = _base_openxml_details(path, preflight, macro_payload)
    details["external_links_detected"] = archive_external_links
    if preflight_limits:
        unreadable = any(
            item["code"] == "unreadable_openxml_archive" for item in preflight_limits
        )
        return details, "failed" if unreadable else "rejected", preflight_limits, 0
    macro_capable = path.suffix.lower() == ".xlsm"
    try:
        formulas = load_workbook(
            path,
            data_only=False,
            read_only=False,
            keep_links=False,
            keep_vba=False,
        )
        cached = load_workbook(
            path,
            data_only=True,
            read_only=False,
            keep_links=False,
            keep_vba=False,
        )
    except Exception:
        return (
            details,
            "failed",
            [
                _limit(
                    "unreadable_openxml_workbook",
                    "Книга Excel не прочитана безопасным локальным парсером.",
                    path.name,
                )
            ],
            0,
        )

    sheets: list[dict[str, Any]] = []
    hidden_sheets: list[str] = []
    merged_cells: list[str] = []
    missing_cached: list[str] = []
    cache_mismatches: list[dict[str, Any]] = []
    limitations: list[dict[str, Any]] = []
    external_links = archive_external_links
    formula_count = 0
    semantic_row_count = 0
    try:
        details["defined_names"] = _defined_names(formulas)
        calculation = getattr(formulas, "calculation", None)
        details["calculation_mode"] = getattr(calculation, "calcMode", None)
        details["calculation"] = {
            key: getattr(calculation, key, None)
            for key in (
                "calcMode",
                "fullCalcOnLoad",
                "forceFullCalc",
                "calcOnSave",
                "iterate",
            )
        }
        for sheet in formulas.worksheets:
            cached_sheet = cached[sheet.title]
            if sheet.sheet_state != "visible":
                hidden_sheets.append(sheet.title)
            merged = [str(cell_range) for cell_range in sheet.merged_cells.ranges]
            merged_cells.extend(f"{sheet.title}!{cell_range}" for cell_range in merged)
            materialised = sorted(
                (
                    cell
                    for cell in sheet._cells.values()
                    if cell.value is not None
                ),
                key=lambda cell: (cell.row, cell.column),
            )
            scalar_values = {
                cell.coordinate.upper(): cell.value
                for cell in materialised
                if cell.data_type != "f" and type(cell.value) in (int, float)
            }
            formula_values = {
                cell.coordinate.upper(): str(cell.value)
                for cell in materialised
                if cell.data_type == "f"
            }
            computed_values, formula_errors = _compute_sheet_formulas(
                formula_values, scalar_values
            )
            formula_count += len(formula_values)
            tables = _table_inventory(sheet)
            cells: list[dict[str, Any]] = []
            occupied: dict[tuple[int, int], Any] = {}
            formats: dict[str, list[str]] = {}
            for cell in materialised:
                locator = f"{sheet.title}!{cell.coordinate}"
                formula = str(cell.value) if cell.data_type == "f" else None
                cached_value = cached_sheet[cell.coordinate].value if formula else cell.value
                computed = computed_values.get(cell.coordinate.upper()) if formula else None
                if formula:
                    if cached_value is None:
                        missing_cached.append(locator)
                    if any(marker in formula for marker in ("!", "[", "]")):
                        external_links = True
                    error = formula_errors.get(cell.coordinate.upper())
                    if error is not None:
                        code = (
                            "formula_cycle"
                            if isinstance(error, FormulaCycle)
                            else "unsupported_formula"
                        )
                        limitations.append(
                            _limit(
                                code,
                                (
                                    "Формула не вычислялась из-за циклической зависимости."
                                    if code == "formula_cycle"
                                    else "Формула не вычислялась: она вне безопасного локального поднабора."
                                ),
                                locator,
                                reason=str(error),
                                formula=formula,
                            )
                        )
                    if (
                        type(cached_value) in (int, float)
                        and type(computed) in (int, float)
                    ):
                        cached_decimal = _decimal_number(cached_value)
                        computed_decimal = _decimal_number(computed)
                        if cached_decimal != computed_decimal:
                            cache_mismatches.append(
                                {
                                    "locator": locator,
                                    "formula": formula,
                                    "cached_value": cached_value,
                                    "computed_safe_value": computed,
                                    "raw_difference": _normalise_number(
                                        computed_decimal - cached_decimal
                                    ),
                                }
                            )
                semantic_value = _json_scalar(
                    cached_value
                    if formula and cached_value is not None
                    else computed
                    if formula and computed is not None
                    else formula
                    if formula
                    else cell.value
                )
                occupied[(cell.row, cell.column)] = semantic_value
                formats.setdefault(cell.number_format, []).append(cell.coordinate)
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "value": None if formula else _json_scalar(cell.value),
                        "formula": formula,
                        "cached_value": _json_scalar(cached_value),
                        "computed_safe_value": computed,
                        "number_format": cell.number_format,
                        "style_id": cell.style_id,
                        "evidence": {"locator": locator},
                    }
                )
            semantic_tables = _semantic_tables(
                sheet_name=sheet.title,
                occupied=occupied,
                locators=lambda row, column, title=sheet.title: (
                    f"{title}!{get_column_letter(column)}{row}"
                ),
                explicit_tables=tables,
                formulas=formula_values,
            )
            semantic_row_count += sum(len(table["rows"]) for table in semantic_tables)
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "hidden_rows": sorted(
                        index
                        for index, dimension in sheet.row_dimensions.items()
                        if dimension.hidden
                    ),
                    "hidden_columns": sorted(
                        key
                        for key, dimension in sheet.column_dimensions.items()
                        if dimension.hidden
                    ),
                    "auto_filter": sheet.auto_filter.ref,
                    "merged_cells": merged,
                    "tables": tables,
                    "number_formats": formats,
                    "semantic_tables": semantic_tables,
                    "cells": cells,
                }
            )
    finally:
        formulas.close()
        cached.close()

    if missing_cached:
        limitations.append(
            {
                "code": "missing_cached_formula_values",
                "message": "Для части формул в книге отсутствуют сохранённые вычисленные значения.",
                "evidence": [{"locator": locator} for locator in missing_cached],
                "details": {"count": len(missing_cached)},
            }
        )
    if external_links:
        limitations.append(
            {
                "code": "external_workbook_links",
                "message": "Обнаружены внешние связи; они не загружались и не вычислялись.",
                "evidence": [{"locator": path.name}],
            }
        )
    if macro_capable or macro_payload:
        limitations.append(
            {
                "code": "macro_capable_workbook",
                "message": "Книга допускает макросы; макросы не загружались и не исполнялись.",
                "evidence": [{"locator": path.name}],
            }
        )
    details.update(
        {
            "sheets": sheets,
            "hidden_sheets": hidden_sheets,
            "merged_cells": merged_cells,
            "macro_capable": macro_capable,
            "vba_payload_present": macro_payload,
            "macros_executed": False,
            "external_links_detected": external_links,
            "missing_cached_formula_values": missing_cached,
            "formula_cache_mismatches": cache_mismatches,
            "semantic_row_count": semantic_row_count,
            "formula_count": formula_count,
        }
    )
    return (
        details,
        "partial" if limitations else "reliable",
        limitations,
        semantic_row_count + formula_count,
    )


def _xls_format(
    workbook: Any, sheet: Any, row: int, column: int
) -> tuple[int | None, str | None]:
    try:
        xf_index = sheet.cell_xf_index(row, column)
        xf = workbook.xf_list[xf_index]
        number_format = workbook.format_map.get(xf.format_key)
        return xf_index, getattr(number_format, "format_str", None)
    except (AttributeError, IndexError, KeyError):
        return None, None


def _xls_names(workbook: Any) -> list[dict[str, Any]]:
    return [
        {
            "name": getattr(item, "name", None),
            "target": getattr(item, "formula_text", None),
            "scope": getattr(item, "scope", None),
            "hidden": bool(getattr(item, "hidden", False)),
        }
        for item in getattr(workbook, "name_obj_list", [])
    ]


def extract_xls(path: Path) -> tuple[dict[str, Any], str, list[dict[str, Any]], int]:
    size = path.stat().st_size
    base_details: dict[str, Any] = {
        "format": "xls",
        "read_only": True,
        "file_size_bytes": size,
        "sheets": [],
        "defined_names": [],
    }
    if size > MAX_EXCEL_FILE_BYTES:
        return (
            base_details,
            "rejected",
            [
                _limit(
                    "excel_file_size_limit",
                    "Размер книги превышает защитный лимит; книга не открывалась.",
                    path.name,
                    actual_bytes=size,
                    limit_bytes=MAX_EXCEL_FILE_BYTES,
                )
            ],
            0,
        )
    try:
        workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=True)
    except Exception:
        return (
            base_details,
            "failed",
            [
                _limit(
                    "unreadable_legacy_xls",
                    "Файл .xls не прочитан библиотекой xlrd в режиме только чтения.",
                    path.name,
                )
            ],
            0,
        )
    try:
        if workbook.nsheets > MAX_WORKBOOK_SHEETS:
            return (
                base_details,
                "rejected",
                [
                    _limit(
                        "excel_sheet_count_limit",
                        "Число листов .xls превышает защитный лимит.",
                        path.name,
                        actual_sheets=workbook.nsheets,
                        limit_sheets=MAX_WORKBOOK_SHEETS,
                    )
                ],
                0,
            )
        dimensions = [(sheet.nrows, sheet.ncols) for sheet in workbook.sheets()]
        declared_cells = sum(rows * columns for rows, columns in dimensions)
        if declared_cells > MAX_WORKBOOK_CELLS:
            return (
                {**base_details, "declared_cell_count": declared_cells},
                "rejected",
                [
                    _limit(
                        "excel_cell_count_limit",
                        "Объявленное число ячеек .xls превышает защитный лимит.",
                        path.name,
                        actual_cells=declared_cells,
                        limit_cells=MAX_WORKBOOK_CELLS,
                    )
                ],
                0,
            )
        sheets: list[dict[str, Any]] = []
        semantic_row_count = 0
        for sheet in workbook.sheets():
            cells: list[dict[str, Any]] = []
            occupied: dict[tuple[int, int], Any] = {}
            formats: dict[str, list[str]] = {}
            for row_index in range(sheet.nrows):
                for column_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, column_index)
                    if value in (None, ""):
                        continue
                    row = row_index + 1
                    column = column_index + 1
                    locator = f"{sheet.name}!R{row}C{column}"
                    occupied[(row, column)] = value
                    xf_index, number_format = _xls_format(
                        workbook, sheet, row_index, column_index
                    )
                    if number_format:
                        formats.setdefault(number_format, []).append(f"R{row}C{column}")
                    cells.append(
                        {
                            "row": row,
                            "column": column,
                            "cached_value": value,
                            "formula": None,
                            "xf_index": xf_index,
                            "number_format": number_format,
                            "evidence": {"locator": locator},
                        }
                    )
            semantic_tables = _semantic_tables(
                sheet_name=sheet.name,
                occupied=occupied,
                locators=lambda row, column, title=sheet.name: f"{title}!R{row}C{column}",
                explicit_tables=[],
            )
            semantic_row_count += sum(len(table["rows"]) for table in semantic_tables)
            sheets.append(
                {
                    "name": sheet.name,
                    "state": "hidden" if getattr(sheet, "visibility", 0) else "visible",
                    "rows": sheet.nrows,
                    "columns": sheet.ncols,
                    "hidden_rows": sorted(
                        index + 1
                        for index, info in getattr(sheet, "rowinfo_map", {}).items()
                        if getattr(info, "hidden", False)
                    ),
                    "hidden_columns": sorted(
                        index + 1
                        for index, info in getattr(sheet, "colinfo_map", {}).items()
                        if getattr(info, "hidden", False)
                    ),
                    "auto_filter": None,
                    "tables": [],
                    "number_formats": formats,
                    "semantic_tables": semantic_tables,
                    "cells": cells,
                }
            )
        limitations = [
            {
                "code": "legacy_xls_formula_metadata_unavailable",
                "message": "Для .xls xlrd возвращает сохранённые значения без надёжного текста формул.",
                "evidence": [{"locator": path.name}],
            }
        ]
        return (
            {
                **base_details,
                "declared_cell_count": declared_cells,
                "defined_names": _xls_names(workbook),
                "sheets": sheets,
                "semantic_row_count": semantic_row_count,
                "formula_count": 0,
            },
            "partial",
            limitations,
            semantic_row_count,
        )
    finally:
        workbook.release_resources()

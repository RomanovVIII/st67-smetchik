from __future__ import annotations

import hashlib
import io
import json
import math
import os
import re
import stat
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from excel_extract import _ooxml_preflight


CORRECTION_SCHEMA_PATH = (
    Path(__file__).resolve().parents[1] / "assets" / "smetchik.corrections.v1.schema.json"
)
MAX_CORRECTIONS_JSON_BYTES = 1024 * 1024
MAX_CORRECTIONS = 1_000
MAX_JSON_DEPTH = 24
MAX_JSON_NODES = 20_000
MAX_REASON_CHARACTERS = 2_000
MAX_VALUE_STRING_CHARACTERS = 8_192
MAX_LOCATOR_CHARACTERS = 1_024
MAX_OUTPUT_BYTES = 200 * 1024 * 1024
MAX_SOURCE_BYTES = 100 * 1024 * 1024
MAX_XML_SOURCE_BYTES = 20 * 1024 * 1024
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xls", ".pdf", ".xml", ".gge", ".mge"}
OOXML_SUFFIXES = {".xlsx", ".xlsm"}
SIDECAR_KINDS = {
    ".pdf": "pdf_sidecar",
    ".xml": "xml_sidecar",
    ".gge": "xml_sidecar",
    ".mge": "xml_sidecar",
    ".xls": "xls_sidecar",
}
ID_PATTERN = re.compile(r"^[A-Za-z0-9_.:-]{1,100}$")
SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")
CELL_PATTERN = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}$")


class CorrectionError(Exception):
    """A safe, user-facing rejection of a correction request."""


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _regular_input(path: Path, *, label: str, size_limit: int) -> os.stat_result:
    try:
        metadata = path.lstat()
    except OSError as error:
        raise CorrectionError(f"{label} does not exist or cannot be inspected") from error
    if stat.S_ISLNK(metadata.st_mode):
        raise CorrectionError(f"{label} symlink is forbidden")
    if not stat.S_ISREG(metadata.st_mode):
        raise CorrectionError(f"{label} must be a regular file")
    if metadata.st_size > size_limit:
        raise CorrectionError(f"{label} size limit exceeded before reading")
    return metadata


def _resolved_candidate(path: Path) -> Path:
    try:
        parent = path.parent.resolve(strict=True)
    except OSError as error:
        raise CorrectionError("artifact parent does not exist or cannot be inspected") from error
    return parent / path.name


def _validate_artifact_path(
    path: Path,
    *,
    label: str,
    forbidden: set[Path],
) -> Path:
    try:
        parent_metadata = path.parent.lstat()
    except OSError as error:
        raise CorrectionError(f"{label} parent does not exist or cannot be inspected") from error
    if stat.S_ISLNK(parent_metadata.st_mode):
        raise CorrectionError(f"{label} parent symlink is forbidden")
    if not stat.S_ISDIR(parent_metadata.st_mode):
        raise CorrectionError(f"{label} parent must be a directory")
    candidate = _resolved_candidate(path)
    if candidate in forbidden:
        raise CorrectionError(f"{label} must not alias the source or corrections input")
    try:
        path.lstat()
    except FileNotFoundError:
        return candidate
    except OSError as error:
        raise CorrectionError(f"{label} cannot be inspected") from error
    raise CorrectionError(f"{label} already exists; overwrite is forbidden")


def _reject_non_finite(token: str) -> None:
    raise CorrectionError(f"non-finite JSON number is forbidden: {token}")


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    value: dict[str, Any] = {}
    for key, item in pairs:
        if key in value:
            raise CorrectionError(f"duplicate JSON property is forbidden: {key}")
        value[key] = item
    return value


def _validate_json_shape(value: Any) -> None:
    nodes = 0
    stack: list[tuple[Any, int]] = [(value, 1)]
    while stack:
        item, depth = stack.pop()
        nodes += 1
        if nodes > MAX_JSON_NODES:
            raise CorrectionError("corrections JSON node limit exceeded")
        if depth > MAX_JSON_DEPTH:
            raise CorrectionError("corrections JSON depth limit exceeded")
        if isinstance(item, dict):
            stack.extend((nested, depth + 1) for nested in item.values())
        elif isinstance(item, list):
            stack.extend((nested, depth + 1) for nested in item)


def _is_scalar(value: Any) -> bool:
    if value is None or isinstance(value, (str, bool, int)):
        return True
    return isinstance(value, float) and math.isfinite(value)


def _validate_scalar(value: Any, *, field: str) -> None:
    if not _is_scalar(value):
        raise CorrectionError(f"{field} must be a finite JSON scalar")
    if isinstance(value, str) and len(value) > MAX_VALUE_STRING_CHARACTERS:
        raise CorrectionError(f"{field} string limit exceeded")


def _validate_corrections_document(value: Any) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise CorrectionError("corrections JSON must contain an object")
    expected_top = {"schema_version", "source_sha256", "corrections"}
    if set(value) != expected_top:
        missing = sorted(expected_top - set(value))
        extra = sorted(set(value) - expected_top)
        if missing:
            raise CorrectionError("missing required correction property: " + ", ".join(missing))
        raise CorrectionError("unexpected correction property: " + ", ".join(extra))
    if value["schema_version"] != "smetchik.corrections.v1":
        raise CorrectionError("unsupported corrections schema_version")
    source_hash = value["source_sha256"]
    if not isinstance(source_hash, str) or SHA256_PATTERN.fullmatch(source_hash) is None:
        raise CorrectionError("source_sha256 must be a lowercase SHA-256 value")
    corrections = value["corrections"]
    if not isinstance(corrections, list) or not corrections:
        raise CorrectionError("corrections must be a non-empty array")
    if len(corrections) > MAX_CORRECTIONS:
        raise CorrectionError("corrections count limit exceeded")
    seen_ids: set[str] = set()
    for index, correction in enumerate(corrections):
        prefix = f"corrections[{index}]"
        if not isinstance(correction, dict):
            raise CorrectionError(f"{prefix} must be an object")
        expected_fields = {
            "id",
            "status",
            "target",
            "expected_old",
            "replacement",
            "reason",
        }
        if set(correction) != expected_fields:
            missing = sorted(expected_fields - set(correction))
            extra = sorted(set(correction) - expected_fields)
            if missing:
                raise CorrectionError(f"{prefix} missing required property: " + ", ".join(missing))
            raise CorrectionError(f"{prefix} unexpected property: " + ", ".join(extra))
        correction_id = correction["id"]
        if not isinstance(correction_id, str) or ID_PATTERN.fullmatch(correction_id) is None:
            raise CorrectionError(f"{prefix}.id is invalid")
        if correction_id in seen_ids:
            raise CorrectionError(f"duplicate correction id: {correction_id}")
        seen_ids.add(correction_id)
        if correction["status"] not in {"confirmed", "disputed"}:
            raise CorrectionError(f"{prefix}.status must be confirmed or disputed")
        if not isinstance(correction["target"], dict):
            raise CorrectionError(f"{prefix}.target must be an object")
        _validate_scalar(correction["expected_old"], field=f"{prefix}.expected_old")
        _validate_scalar(correction["replacement"], field=f"{prefix}.replacement")
        reason = correction["reason"]
        if not isinstance(reason, str) or not reason.strip():
            raise CorrectionError(f"{prefix}.reason must be a non-empty string")
        if len(reason) > MAX_REASON_CHARACTERS:
            raise CorrectionError(f"{prefix}.reason string limit exceeded")
    return value


def _read_corrections(path: Path) -> dict[str, Any]:
    try:
        raw = path.read_bytes()
        text = raw.decode("utf-8")
        value = json.loads(
            text,
            parse_constant=_reject_non_finite,
            object_pairs_hook=_reject_duplicate_keys,
        )
    except CorrectionError:
        raise
    except (OSError, UnicodeError, json.JSONDecodeError, RecursionError) as error:
        raise CorrectionError("corrections JSON is unreadable or invalid") from error
    _validate_json_shape(value)
    return _validate_corrections_document(value)


def _validate_excel_target(target: dict[str, Any], *, prefix: str) -> tuple[str, str]:
    if set(target) != {"sheet", "cell"}:
        raise CorrectionError(f"{prefix}.target must contain only sheet and cell")
    sheet = target["sheet"]
    coordinate = target["cell"]
    if not isinstance(sheet, str) or not sheet or len(sheet) > 31:
        raise CorrectionError(f"{prefix}.target.sheet is invalid")
    if not isinstance(coordinate, str) or CELL_PATTERN.fullmatch(coordinate) is None:
        raise CorrectionError(f"{prefix}.target.cell is invalid")
    column, row = coordinate_from_string(coordinate.upper())
    if column_index_from_string(column) > 16_384 or row > 1_048_576:
        raise CorrectionError(f"{prefix}.target.cell is outside the Excel grid")
    return sheet, f"{column}{row}"


def _validate_sidecar_target(target: dict[str, Any], *, prefix: str) -> str:
    if set(target) != {"locator"}:
        raise CorrectionError(f"{prefix}.target must contain only locator")
    locator = target["locator"]
    if not isinstance(locator, str) or not locator.strip():
        raise CorrectionError(f"{prefix}.target.locator is invalid")
    if len(locator) > MAX_LOCATOR_CHARACTERS:
        raise CorrectionError(f"{prefix}.target.locator string limit exceeded")
    return locator


def _safe_json_bytes(value: dict[str, Any]) -> bytes:
    try:
        serialized = json.dumps(
            value,
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
            allow_nan=False,
        )
    except (TypeError, ValueError) as error:
        raise CorrectionError("artifact JSON could not be serialized safely") from error
    return (serialized + "\n").encode("utf-8")


def _normalized_cell_value(value: Any) -> Any:
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return value


def _values_equal(actual: Any, expected: Any) -> bool:
    actual = _normalized_cell_value(actual)
    if isinstance(actual, bool) or isinstance(expected, bool):
        return type(actual) is type(expected) and actual == expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        try:
            return Decimal(str(actual)) == Decimal(str(expected))
        except InvalidOperation:
            return False
    return type(actual) is type(expected) and actual == expected


def _forbidden_active_replacement(value: Any) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@"))


def _xlsx_artifact(
    source: Path,
    spec: dict[str, Any],
    source_hash: str,
) -> tuple[bytes, dict[str, Any], list[dict[str, str]], str]:
    _preflight, preflight_limits, _macro_payload, external_links = _ooxml_preflight(source)
    if preflight_limits:
        codes = ", ".join(str(item.get("code") or "unknown") for item in preflight_limits)
        raise CorrectionError(f"source spreadsheet failed safety preflight: {codes}")
    keep_vba = source.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(
            source,
            data_only=False,
            read_only=False,
            keep_vba=keep_vba,
            keep_links=True,
        )
    except Exception as error:
        raise CorrectionError("source spreadsheet cannot be opened safely") from error

    validated: list[tuple[dict[str, Any], str, str, Any]] = []
    seen_targets: set[tuple[str, str]] = set()
    try:
        for index, correction in enumerate(spec["corrections"]):
            prefix = f"corrections[{index}]"
            sheet_name, coordinate = _validate_excel_target(correction["target"], prefix=prefix)
            target_key = (sheet_name, coordinate)
            if target_key in seen_targets:
                raise CorrectionError(f"duplicate correction target: {sheet_name}!{coordinate}")
            seen_targets.add(target_key)
            if sheet_name not in workbook.sheetnames:
                raise CorrectionError(f"correction target sheet is absent: {sheet_name}")
            cell = workbook[sheet_name][coordinate]
            actual = _normalized_cell_value(cell.value)
            if not _values_equal(actual, correction["expected_old"]):
                raise CorrectionError(
                    f"stale expected_old for correction {correction['id']} at {sheet_name}!{coordinate}"
                )
            if correction["status"] == "confirmed" and _forbidden_active_replacement(
                correction["replacement"]
            ):
                raise CorrectionError("formula or active external-link replacement is forbidden in v1")
            replacement = correction["replacement"]
            if correction["status"] == "confirmed" and isinstance(replacement, str):
                if ILLEGAL_CHARACTERS_RE.search(replacement):
                    raise CorrectionError("replacement contains characters forbidden by OOXML")
            validated.append((correction, sheet_name, coordinate, actual))

        changes: list[dict[str, Any]] = []
        for correction, sheet_name, coordinate, actual in validated:
            cell = workbook[sheet_name][coordinate]
            target = {"sheet": sheet_name, "cell": coordinate}
            if correction["status"] == "confirmed":
                cell.value = correction["replacement"]
                action = "replaced"
            else:
                proposal = json.dumps(
                    correction["replacement"], ensure_ascii=False, allow_nan=False
                )
                annotation = (
                    f"Сметчик — спорное предложение {correction['id']}.\n"
                    f"Предлагаемое значение: {proposal}\n"
                    f"Основание: {correction['reason']}"
                )
                existing = cell.comment.text + "\n\n" if cell.comment is not None else ""
                comment_text = existing + annotation
                if len(comment_text) > 32_000 or ILLEGAL_CHARACTERS_RE.search(comment_text):
                    raise CorrectionError("annotation comment exceeds safe OOXML limits")
                cell.comment = Comment(comment_text, "Сметчик")
                action = "annotated"
            changes.append(
                {
                    "id": correction["id"],
                    "action": action,
                    "target": target,
                    "old_value": actual,
                    "new_value": correction["replacement"],
                    "reason": correction["reason"],
                }
            )

        stream = io.BytesIO()
        workbook.save(stream)
        payload = stream.getvalue()
    except CorrectionError:
        raise
    except Exception as error:
        raise CorrectionError("corrected spreadsheet could not be serialized safely") from error
    finally:
        workbook.close()
    if len(payload) > MAX_OUTPUT_BYTES:
        raise CorrectionError("corrected spreadsheet output size limit exceeded")
    limitations: list[dict[str, str]] = []
    if keep_vba:
        limitations.append(
            {
                "code": "xlsm_vba_preserved_without_execution",
                "message": "VBA payload was preserved with keep_vba; macros were not executed or verified.",
            }
        )
    if external_links:
        limitations.append(
            {
                "code": "external_links_preserved_not_resolved",
                "message": "Existing external links were preserved but never opened, refreshed, or verified.",
            }
        )
    artifact_kind = "xlsm_copy" if keep_vba else "xlsx_copy"
    log = {
        "schema_version": "smetchik.changelog.v1",
        "artifact_kind": artifact_kind,
        "source_sha256_before": source_hash,
        "source_sha256_after": source_hash,
        "output_sha256": _sha256_bytes(payload),
        "macros_executed": False,
        "changes": changes,
        "limitations": limitations,
    }
    return payload, log, limitations, artifact_kind


def _sidecar_artifact(
    source: Path,
    spec: dict[str, Any],
    source_hash: str,
) -> tuple[bytes, dict[str, Any], list[dict[str, str]], str]:
    suffix = source.suffix.lower()
    artifact_kind = SIDECAR_KINDS[suffix]
    seen_targets: set[str] = set()
    proposals: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    for index, correction in enumerate(spec["corrections"]):
        locator = _validate_sidecar_target(correction["target"], prefix=f"corrections[{index}]")
        if locator in seen_targets:
            raise CorrectionError(f"duplicate correction target locator: {locator}")
        seen_targets.add(locator)
        proposal = {
            "id": correction["id"],
            "status": correction["status"],
            "target": {"locator": locator},
            "expected_old": correction["expected_old"],
            "replacement": correction["replacement"],
            "reason": correction["reason"],
            "applied": False,
        }
        proposals.append(proposal)
        changes.append(
            {
                "id": correction["id"],
                "action": "proposed_sidecar_only",
                "target": {"locator": locator},
                "old_value": correction["expected_old"],
                "new_value": correction["replacement"],
                "reason": correction["reason"],
            }
        )
    limitations = [
        {
            "code": (
                "legacy_xls_editing_unsupported_sidecar_only"
                if suffix == ".xls"
                else "source_format_not_modified_sidecar_only"
            ),
            "message": "The source format was not edited; corrections are proposals in a sidecar.",
        }
    ]
    artifact = {
        "schema_version": "smetchik.correction-sidecar.v1",
        "artifact_kind": artifact_kind,
        "source_sha256": source_hash,
        "proposals": proposals,
        "limitations": limitations,
    }
    payload = _safe_json_bytes(artifact)
    log = {
        "schema_version": "smetchik.changelog.v1",
        "artifact_kind": artifact_kind,
        "source_sha256_before": source_hash,
        "source_sha256_after": source_hash,
        "output_sha256": _sha256_bytes(payload),
        "macros_executed": False,
        "changes": changes,
        "limitations": limitations,
    }
    return payload, log, limitations, artifact_kind


def _unlink_owned(path: Path, metadata: os.stat_result) -> None:
    try:
        current = path.lstat()
        if (
            stat.S_ISREG(current.st_mode)
            and current.st_dev == metadata.st_dev
            and current.st_ino == metadata.st_ino
        ):
            path.unlink()
    except FileNotFoundError:
        pass


def _exclusive_write(path: Path, payload: bytes) -> os.stat_result:
    descriptor: int | None = None
    metadata: os.stat_result | None = None
    try:
        descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        metadata = os.fstat(descriptor)
        with os.fdopen(descriptor, "wb", closefd=True) as stream:
            descriptor = None
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        return metadata
    except FileExistsError as error:
        raise CorrectionError("artifact already exists; overwrite is forbidden") from error
    except OSError as error:
        if descriptor is not None:
            os.close(descriptor)
        if metadata is not None:
            _unlink_owned(path, metadata)
        raise CorrectionError("artifact could not be created safely") from error


def _publish_pair(
    output: Path,
    output_payload: bytes,
    changelog: Path,
    changelog_payload: bytes,
) -> None:
    output_metadata = _exclusive_write(output, output_payload)
    try:
        _exclusive_write(changelog, changelog_payload)
    except CorrectionError:
        _unlink_owned(output, output_metadata)
        raise


def create_corrected_artifacts(
    source: Path,
    corrections_json: Path,
    output: Path,
    changelog: Path,
) -> dict[str, Any]:
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise CorrectionError("source format is unsupported for correction artifacts")
    source_limit = MAX_XML_SOURCE_BYTES if suffix in {".xml", ".gge", ".mge"} else MAX_SOURCE_BYTES
    _regular_input(source, label="source", size_limit=source_limit)
    _regular_input(
        corrections_json,
        label="corrections JSON",
        size_limit=MAX_CORRECTIONS_JSON_BYTES,
    )
    source_resolved = source.resolve(strict=True)
    corrections_resolved = corrections_json.resolve(strict=True)
    forbidden = {source_resolved, corrections_resolved}
    output_resolved = _validate_artifact_path(output, label="output", forbidden=forbidden)
    changelog_resolved = _validate_artifact_path(
        changelog, label="changelog", forbidden=forbidden
    )
    if output_resolved == changelog_resolved:
        raise CorrectionError("output and changelog paths must be distinct")
    if changelog.suffix.lower() != ".json":
        raise CorrectionError("changelog must use the .json extension")
    if suffix in OOXML_SUFFIXES:
        if output.suffix.lower() != suffix:
            raise CorrectionError("corrected OOXML output must preserve the source extension")
    elif output.suffix.lower() != ".json":
        raise CorrectionError("sidecar output must use the .json extension")

    source_hash_before = _sha256_file(source)
    spec = _read_corrections(corrections_json)
    if spec["source_sha256"] != source_hash_before:
        raise CorrectionError("source hash does not match corrections JSON")
    if suffix in OOXML_SUFFIXES:
        output_payload, log, limitations, artifact_kind = _xlsx_artifact(
            source, spec, source_hash_before
        )
    else:
        output_payload, log, limitations, artifact_kind = _sidecar_artifact(
            source, spec, source_hash_before
        )
    source_hash_after = _sha256_file(source)
    if source_hash_after != source_hash_before:
        raise CorrectionError("source changed during correction; no artifact was published")
    log["source_sha256_after"] = source_hash_after
    changelog_payload = _safe_json_bytes(log)
    _publish_pair(output, output_payload, changelog, changelog_payload)
    return {
        "schema_version": "smetchik.correction-run.v1",
        "artifact_kind": artifact_kind,
        "source_sha256": source_hash_before,
        "output_sha256": _sha256_bytes(output_payload),
        "change_count": len(spec["corrections"]),
        "output": output.name,
        "changelog": changelog.name,
        "limitations": limitations,
    }

